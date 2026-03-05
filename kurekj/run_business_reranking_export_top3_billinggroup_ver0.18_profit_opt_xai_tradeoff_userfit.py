# run_business_reranking_export_top3_billinggroup_ver0.17_profit_opt_xai.py
from __future__ import annotations

from dataclasses import asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import numpy as np
import pandas as pd
from pandas.api.types import DatetimeTZDtype
from tqdm.auto import tqdm

from src.v2.api.services.business_reranking6 import (  # type: ignore
    Candidate,
    RerankParams,
    business_score,
    business_score_breakdown,
    greedy_rerank,
)

# =========================
# CONFIG
# =========================

RECS_PATH = Path("recommendations_output.parquet")
OUT_XLSX_BASENAME = "compare_top3"

TOPK = 3
# P_H = 0.0
P_H = 0.047206591673

API_MODEL_SCORE_MISSING_SENTINEL = -1.0

DEFAULT_CONTRACT_TYPE = "flat"
DEFAULT_CAP_RATIO = 0.0
DEFAULT_M = 0.0
DEFAULT_V = 0

# -------------------------
# Feature normalization (IMPORTANT)
# -------------------------
# The score expects:
#   r,g in [0,1], cap_ratio in [0,inf), m in [-1,1], v in {0,1}.
# This script builds these ranges explicitly.

FAIRNESS_MODE = "budget_norm_log1p"  # {'none','budget_norm_log1p'}
FAIRNESS_BUDGET_BY_CONTRACT = True   # normalize budgets separately for cpl vs flat
FAIRNESS_CLIP = (-1.0, 1.0)

VIP_MODE = "p90_effective_price"     # {'none','binary_from_lead_limit','p90_effective_price'}
VIP_BY_CONTRACT = True              # compute VIP thresholds separately for cpl vs flat

# --- daily_snapshot.offer_performance_metrics snapshot inputs ---
OFFERS_OPM_DATA_PATH = Path("daily_snapshot_offer_performance_metrics_data_20260120_174227.xlsx")
OFFERS_OPM_META_PATH = Path("daily_snapshot_offer_performance_metrics_meta_20260120_174227.xlsx")

OFFERS_OPM_DAY_SHIFT_DAYS = 0  # int
OFFERS_OPM_ASOF_DIRECTION = "backward"  # {'backward','forward','nearest'}

CONTRACT_TYPE_CPL = "cpl"   # settlement_type==0
CONTRACT_TYPE_FLAT = "flat" # settlement_type==1 (also DEFAULT_CONTRACT_TYPE)

RERANK_PARAMS_DICT = dict(
    gamma=0.114683695449,
    mu=1.818253964065,
    nu=0.562795525262,
    rho=0.775322715304,
    delta=1.541089556366,
    lambda_=1.620422743819,
)

# --- Monetization / profit proxy (for choosing pg vs kurekj) ---
POSITION_WEIGHT_MODE = "dcg"  # {'dcg','equal','custom'}
CUSTOM_POSITION_WEIGHTS = {1: 1.0, 2: 0.7, 3: 0.5}  # used only if POSITION_WEIGHT_MODE='custom'

PROFIT_PROB_COL = "q_api"  # {'q_api','q_final'} etc.
BILLABLE_CAP_THRESHOLD = 1.0

# =========================
# TRADE-OFF: revenue vs score
# =========================
# Pytanie: „przychód można zwiększyć, ale jakim kosztem jakości/score?”
TRADEOFF_BINS = 10                  # decyle (lub mniejsza liczba)
TRADEOFF_BIN_MODE = "per_algo"      # {'per_algo','global'}
TRADEOFF_BASELINE_ALGO = "baseline" # odniesienie w tabeli „koszt” (quadrants)
TRADEOFF_QUALITY_COL = "q_prob_wmean"  # co traktujemy jako „score jakościowy” do kosztu
TRADEOFF_EPS = 1e-9                 # tolerancja na tie w różnicach

# =========================
# USER FIT: distance from user/model ranking (complaint risk proxy)
# =========================
# Idea:
#   Jeśli re-ranking biznesowy promuje ofertę, która w "user/model baseline" (api_model_score)
#   była daleko w rankingu, użytkownik może odebrać to jako gorszą rekomendację -> większe ryzyko reklamacji.
#
#   Dlatego mierzymy "skąd" (z którego miejsca w baseline) pochodzi rekomendowana oferta.
#   To NIE jest ground truth jakości, ale prosty, czytelny wskaźnik ryzyka.
#
# Kolumny generowane w analizie:
#   - baseline_rank_full (per request_id, wg api_model_score)
#   - n_candidates (ile było kandydatów w request)
#   - orig_rank_top1 (baseline rank rekomendacji TOP-1 po re-rankingu)
#   - user_fit_wmean (0..1, im wyżej tym bliżej baseline)
USER_FIT_ENABLED = True
USER_FIT_SAFE_TOPN = 3                 # np. "bezpieczne": TOP-1 pochodzi z baseline TOP-3
USER_FIT_THRESHOLDS = [3, 5, 10, 15]   # raportowane progi dla orig_rank_top1
USER_FIT_TOP1_BINS = [0, 1, 3, 5, 10, 15, 20, 50, np.inf]
USER_FIT_TOP1_BIN_LABELS = ["=1", "2-3", "4-5", "6-10", "11-15", "16-20", "21-50", "51+"]
USER_FIT_QUALITY_COL = "user_fit_wmean"  # alternatywny quality_col do tradeoff_quadrants

LABEL_RECS = "hm-reranking-testgroup-202512.csv"
LABEL_OPM = "daily_snapshot.offer_performance_metrics"
LABEL_FEATURES = "business reranking: zmienne/cechy (q,r,g,m,v,cap_ratio,...)"
LABEL_WEIGHTS = "business reranking: wagi/parametry (gamma,mu,nu,...)"
LABEL_OUTPUTS = "wyniki rerankingu (rank/score)"
LABEL_XAI = "business_score_breakdown (XAI)"
LABEL_OTHER = "inne"

GROUP_BG = {
    LABEL_RECS: "#DCE6F1",      # light blue
    LABEL_OPM: "#E2EFDA",       # light green
    LABEL_FEATURES: "#FFF2CC",  # light yellow
    LABEL_WEIGHTS: "#FCE4D6",   # light orange
    LABEL_OUTPUTS: "#E4DFEC",   # light purple/gray
    LABEL_XAI: "#DAEEF3",       # light cyan
    LABEL_OTHER: "#F2F2F2",     # light gray
}

# =========================

RECS_REQUIRED_COLS = [
    "create_date",
    "uuid",
    "request_id",
    "offer_id",
    "property_id",
    "api_model_score",
    "final_score",
    "gap_score",
    "lead_price_score",
]

OFFERS_OPM_REQUIRED_COLS = [
    "id",
    "day",
    "offer_id",
    "group_id",
    "monthly_budget",
    "estimated_monthly_budget",
    "group_budget",
    "lead_limit",
    "leads",
    "leads_estimation",
    "estimated_perc_realization",
    "lead_price",
    "estimated_lead_price",
    "settlement_type",
]

# =========================
# HELPERS
# =========================


def read_table(path: Path) -> pd.DataFrame:
    """Generic reader for recs input."""
    suffix = path.suffix.lower()
    if suffix in {".parquet", ".pq"}:
        return pd.read_parquet(path)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    if suffix == ".csv":
        return pd.read_csv(path)
    raise ValueError(f"Unsupported input format: {suffix}")


def ensure_required_columns(df: pd.DataFrame, required: list[str], *, df_name: str = "df") -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"{df_name}: missing required columns: {missing}")


def normalize_create_date(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize create_date to tz-naive Europe/Warsaw (for stable day flooring)."""
    out = df.copy()
    dt = pd.to_datetime(out["create_date"], errors="raise")

    if isinstance(dt.dtype, DatetimeTZDtype):
        dt = dt.dt.tz_convert("Europe/Warsaw").dt.tz_localize(None)

    out["create_date"] = dt
    return out


def _col_or_nan(df: pd.DataFrame, col: str) -> pd.Series:
    """Return df[col] if exists else a NaN series (same length)."""
    if col in df.columns:
        return df[col]
    return pd.Series(np.nan, index=df.index)


def robust_minmax_log1p(series: pd.Series, *, by: pd.Series | None = None) -> pd.Series:
    """
    Robust-ish normalization to [0,1] using log1p + min/max within group.

    Returns:
      float series in [0,1] (NaN -> 0).
    """
    x = pd.to_numeric(series, errors="coerce").astype(float)
    x = np.log1p(x.clip(lower=0.0))

    if by is None:
        lo = float(np.nanmin(x.values)) if np.isfinite(x.values).any() else 0.0
        hi = float(np.nanmax(x.values)) if np.isfinite(x.values).any() else 1.0
        denom = hi - lo if hi > lo else 1.0
        out = (x - lo) / denom
        return out.fillna(0.0).clip(0.0, 1.0)

    out = pd.Series(0.0, index=x.index, dtype=float)
    by_key = by.astype(str).fillna("")

    for _, idx in by_key.groupby(by_key).groups.items():
        xs = x.loc[idx]
        finite = np.isfinite(xs.values)
        if not finite.any():
            out.loc[idx] = 0.0
            continue
        lo = float(np.nanmin(xs.values))
        hi = float(np.nanmax(xs.values))
        denom = hi - lo if hi > lo else 1.0
        out.loc[idx] = ((xs - lo) / denom).fillna(0.0).clip(0.0, 1.0)

    return out


def compute_fairness_from_budget(df: pd.DataFrame) -> pd.Series:
    """
    Build m in [-1,1] from budget.

    Default: log1p(budget) -> [0,1] normalized (optionally per contract),
    then center to [-1,1] via 2*x - 1.
    """
    if FAIRNESS_MODE == "none":
        return pd.Series(0.0, index=df.index, dtype=float)

    budget = pd.to_numeric(_col_or_nan(df, "m_budget_raw"), errors="coerce").astype(float)

    by = None
    if FAIRNESS_BUDGET_BY_CONTRACT:
        by = _col_or_nan(df, "contract_type").astype(str).str.lower()

    norm01 = robust_minmax_log1p(budget, by=by)
    m = 2.0 * norm01 - 1.0

    lo, hi = FAIRNESS_CLIP
    return m.clip(lo, hi).fillna(0.0).astype(float)


def compute_effective_lead_price_for_features(df: pd.DataFrame) -> pd.Series:
    """
    Unified price per lead idea used to build VIP flag.
    """
    lead_price = pd.to_numeric(_col_or_nan(df, "lead_price"), errors="coerce")
    est_lead_price = pd.to_numeric(_col_or_nan(df, "estimated_lead_price"), errors="coerce")

    settlement = pd.to_numeric(_col_or_nan(df, "settlement_type"), errors="coerce")
    contract = _col_or_nan(df, "contract_type").astype(str).str.lower()

    if settlement.notna().any():
        is_cpl = settlement.eq(0)
    else:
        is_cpl = contract.eq(CONTRACT_TYPE_CPL)

    eff = lead_price.where(is_cpl, est_lead_price)

    lead_limit = pd.to_numeric(_col_or_nan(df, "lead_limit_raw"), errors="coerce").replace({0: np.nan})
    est_budget = pd.to_numeric(_col_or_nan(df, "estimated_monthly_budget"), errors="coerce")
    monthly_budget = pd.to_numeric(_col_or_nan(df, "monthly_budget"), errors="coerce")
    leads_est = pd.to_numeric(_col_or_nan(df, "leads_estimation"), errors="coerce").replace({0: np.nan})

    budget_any = est_budget.fillna(monthly_budget)
    fallback1 = budget_any / lead_limit
    fallback2 = budget_any / leads_est

    eff = eff.fillna(fallback1).fillna(fallback2)

    return pd.to_numeric(eff, errors="coerce").astype(float)


def compute_vip_flag(df: pd.DataFrame) -> pd.Series:
    """
    Build v in {0,1}.

    Default: v=1 if effective_lead_price is in top 10% (p90) within (optional) contract segment.
    Alternative: v=1 if lead_limit_raw > 0 (legacy behavior).
    """
    if VIP_MODE == "none":
        return pd.Series(0, index=df.index, dtype=int)

    if VIP_MODE == "binary_from_lead_limit":
        lead_limit = pd.to_numeric(_col_or_nan(df, "lead_limit_raw"), errors="coerce").fillna(0.0)
        return (lead_limit > 0).astype(int)

    price = pd.to_numeric(_col_or_nan(df, "effective_lead_price"), errors="coerce")

    by = None
    if VIP_BY_CONTRACT:
        by = _col_or_nan(df, "contract_type").astype(str).str.lower()

    if by is None:
        thr = float(price.quantile(0.9)) if price.notna().any() else np.inf
        return (price >= thr).fillna(False).astype(int)

    out = pd.Series(0, index=df.index, dtype=int)
    by_key = by.astype(str).fillna("")

    for _, idx in by_key.groupby(by_key).groups.items():
        p = price.loc[idx]
        thr = float(p.quantile(0.9)) if p.notna().any() else np.inf
        out.loc[idx] = (p >= thr).fillna(False).astype(int)

    return out


def read_offer_performance_metrics(path: Path) -> pd.DataFrame:
    """Read daily_snapshot.offer_performance_metrics XLSX dump (sheet: 'data')."""
    if not path.exists():
        raise FileNotFoundError(f"offer_performance_metrics data file not found: {path.resolve()}")

    df = pd.read_excel(path, sheet_name="data")
    ensure_required_columns(df, OFFERS_OPM_REQUIRED_COLS, df_name="offer_performance_metrics")

    out = df.copy()
    out["day"] = pd.to_datetime(out["day"], errors="raise").dt.normalize()

    out["offer_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")
    out["group_id"] = pd.to_numeric(out["group_id"], errors="coerce").astype("Int64")
    out["settlement_type"] = pd.to_numeric(out["settlement_type"], errors="coerce").astype("Int64")

    for c in [
        "monthly_budget",
        "estimated_monthly_budget",
        "group_budget",
        "estimated_perc_realization",
        "lead_price",
        "estimated_lead_price",
    ]:
        out[c] = pd.to_numeric(out[c], errors="coerce").astype(float)

    for c in ["lead_limit", "leads", "leads_estimation"]:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    dup = out.duplicated(["offer_id", "day"], keep=False)
    if dup.any():
        sample = out.loc[dup, ["offer_id", "day"]].head(10)
        raise ValueError(
            "offer_performance_metrics: duplicate rows for (offer_id, day). "
            "This breaks merge_asof assumptions. Sample:\n"
            f"{sample}"
        )

    return out


def enrich_with_offer_performance_metrics(
    recs: pd.DataFrame,
    offers_opm: pd.DataFrame,
) -> pd.DataFrame:
    """Enrich recs with daily_snapshot.offer_performance_metrics and build Candidate business features."""
    if OFFERS_OPM_ASOF_DIRECTION not in {"backward", "forward", "nearest"}:
        raise ValueError(
            "OFFERS_OPM_ASOF_DIRECTION must be one of {'backward','forward','nearest'}, "
            f"got: {OFFERS_OPM_ASOF_DIRECTION!r}"
        )

    out = recs.copy()
    out["offer_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")

    out["rec_day"] = out["create_date"].dt.floor("D") + pd.to_timedelta(
        OFFERS_OPM_DAY_SHIFT_DAYS, unit="D"
    )

    out["_orig_row_order"] = out.index.astype(int)

    left = out.sort_values(["rec_day", "offer_id"], kind="mergesort").reset_index(drop=True)
    right = offers_opm.sort_values(["day", "offer_id"], kind="mergesort").reset_index(drop=True)

    merged = pd.merge_asof(
        left,
        right,
        by="offer_id",
        left_on="rec_day",
        right_on="day",
        direction=OFFERS_OPM_ASOF_DIRECTION,
    )

    merged = (
        merged.sort_values("_orig_row_order", kind="mergesort")
        .drop(columns=["_orig_row_order"])
        .reset_index(drop=True)
    )

    settlement_num = pd.to_numeric(merged.get("settlement_type"), errors="coerce")
    merged["contract_type"] = DEFAULT_CONTRACT_TYPE
    merged.loc[settlement_num == 0, "contract_type"] = CONTRACT_TYPE_CPL
    merged.loc[settlement_num == 1, "contract_type"] = CONTRACT_TYPE_FLAT

    merged["group_id"] = pd.to_numeric(merged.get("group_id"), errors="coerce").astype("Int64")
    merged["inv_id"] = merged["group_id"].fillna(merged["offer_id"]).astype("Int64")

    budget = pd.to_numeric(merged.get("monthly_budget"), errors="coerce")
    est_budget = pd.to_numeric(merged.get("estimated_monthly_budget"), errors="coerce")
    merged["m"] = budget.fillna(est_budget).fillna(DEFAULT_M).astype(float)

    merged["v"] = (
        pd.to_numeric(merged.get("lead_limit"), errors="coerce")
        .fillna(0)
        .astype(int)
    )

    cap = pd.to_numeric(merged.get("estimated_perc_realization"), errors="coerce")
    leads = pd.to_numeric(merged.get("leads"), errors="coerce")
    lead_limit = pd.to_numeric(merged.get("lead_limit"), errors="coerce").replace({0: np.nan})
    cap_fallback = leads / lead_limit
    cap = cap.fillna(cap_fallback)

    merged["cap_ratio"] = (
        cap.fillna(DEFAULT_CAP_RATIO)
        .clip(lower=0.0)
        .astype(float)
    )

    total_rows = len(merged)
    matched_rows = int(merged["id"].notna().sum()) if "id" in merged.columns else 0
    pct_matched = matched_rows / total_rows if total_rows else 0.0

    print(
        "offer_performance_metrics merge summary:\n"
        f"  rows_total..................: {total_rows}\n"
        f"  rows_matched_snapshot.......: {matched_rows} ({pct_matched:.1%})\n"
        f"  snapshot_day_min/max........: {offers_opm['day'].min()} / {offers_opm['day'].max()}"
    )

    return merged


def _prepare_q(series: pd.Series) -> pd.Series:
    q = pd.to_numeric(series, errors="raise").fillna(0.0)
    if API_MODEL_SCORE_MISSING_SENTINEL is not None:
        q = q.mask(q == API_MODEL_SCORE_MISSING_SENTINEL, 0.0)
    return q.clip(-1.0, 1.0)


def build_features(df: pd.DataFrame) -> pd.DataFrame:
    """Build all model/reranking features (with correct ranges for business_score)."""
    out = df.copy().reset_index(drop=True)
    out["row_id"] = out.index.astype(int)

    for c in ["api_model_score", "final_score", "gap_score", "lead_price_score"]:
        out[c] = pd.to_numeric(out[c], errors="raise")

    out["q_api"] = _prepare_q(out["api_model_score"])
    out["q_final"] = _prepare_q(out["final_score"])

    out["r"] = out["lead_price_score"].fillna(0.0).clip(0.0, 1.0)
    out["g"] = out["gap_score"].fillna(0.0).clip(0.0, 1.0)

    if "contract_type" in out.columns:
        out["contract_type"] = out["contract_type"].fillna(DEFAULT_CONTRACT_TYPE).astype(str)
    else:
        out["contract_type"] = DEFAULT_CONTRACT_TYPE

    if "cap_ratio" in out.columns:
        out["cap_ratio"] = pd.to_numeric(out["cap_ratio"], errors="coerce").fillna(DEFAULT_CAP_RATIO).clip(lower=0.0)
    else:
        out["cap_ratio"] = float(DEFAULT_CAP_RATIO)

    if "inv_id" in out.columns:
        out["inv_id"] = pd.to_numeric(out["inv_id"], errors="coerce").astype("Int64")
    else:
        out["inv_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")

    if "m" in out.columns:
        out["m_budget_raw"] = pd.to_numeric(out["m"], errors="coerce").fillna(0.0).astype(float)
    else:
        out["m_budget_raw"] = 0.0

    if "v" in out.columns:
        out["lead_limit_raw"] = pd.to_numeric(out["v"], errors="coerce").fillna(0).astype(int)
    else:
        out["lead_limit_raw"] = 0

    out["effective_lead_price"] = compute_effective_lead_price_for_features(out)
    out["m"] = compute_fairness_from_budget(out)
    out["v"] = compute_vip_flag(out)

    return out


# =========================
# XAI / BREAKDOWN HELPERS
# =========================


def _breakdown_to_dict(bd: Any) -> Dict[str, Any]:
    """
    Convert breakdown -> dict robustly:
      - dict -> as is
      - has .to_dict() -> use it
      - dataclass -> asdict
      - object with __dict__ -> use __dict__
    """
    if bd is None:
        return {}
    if isinstance(bd, dict):
        return dict(bd)
    if hasattr(bd, "to_dict") and callable(getattr(bd, "to_dict")):
        try:
            out = bd.to_dict()
            return dict(out) if isinstance(out, dict) else {}
        except Exception:
            pass
    if is_dataclass(bd):
        try:
            return dict(asdict(bd))
        except Exception:
            pass
    if hasattr(bd, "__dict__"):
        try:
            return dict(bd.__dict__)
        except Exception:
            pass
    return {"_breakdown_repr": repr(bd)}


def _call_breakdown(cand: Candidate, params: RerankParams, *, p_h: float, selected: list[Candidate]) -> Any:
    """
    Call business_score_breakdown with best-effort signature compatibility.
    """
    try:
        return business_score_breakdown(cand, params, p_h=p_h, selected=selected)
    except TypeError:
        return business_score_breakdown(cand, params, p_h=p_h)


# =========================
# TOP-K
# =========================



def add_baseline_full_ranks(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds full baseline ranks per request_id based on api_model_score (descending).

    This is used by USER_FIT metric to quantify how far the final (business) recommendation
    moved away from the user/model baseline ranking.

    Adds columns:
      - baseline_rank_full : Int64 (1 = best in baseline)
      - n_candidates       : Int64 (number of candidates in the request)

    Notes:
      - Requires build_features() to have been run (needs row_id).
      - Tie-break uses property_id for deterministic ordering (same as baseline_topk_by()).
    """
    if "row_id" not in df.columns:
        raise ValueError("add_baseline_full_ranks: missing 'row_id' (run build_features() first)")
    for c in ["request_id", "api_model_score", "property_id"]:
        if c not in df.columns:
            raise ValueError(f"add_baseline_full_ranks: missing required column: {c!r}")

    tmp = (
        df.sort_values(
            ["request_id", "api_model_score", "property_id"],
            ascending=[True, False, True],
            kind="mergesort",
        )[["row_id", "request_id"]]
        .copy()
    )

    tmp["baseline_rank_full"] = tmp.groupby("request_id").cumcount() + 1
    tmp["n_candidates"] = tmp.groupby("request_id")["request_id"].transform("size")

    # map back to original order via row_id (stable)
    rank_map = tmp.set_index("row_id")["baseline_rank_full"]
    n_map = tmp.set_index("row_id")["n_candidates"]

    out = df.copy()
    out["baseline_rank_full"] = pd.to_numeric(out["row_id"].map(rank_map), errors="coerce").astype("Int64")
    out["n_candidates"] = pd.to_numeric(out["row_id"].map(n_map), errors="coerce").astype("Int64")
    return out


def baseline_topk_by(df: pd.DataFrame, k: int) -> pd.DataFrame:
    """
    BASELINE = PROVIDED:
      - api_model_score to dostarczony finalny biznesowy score/ranking
      - NIE wolno liczyć business_score() dla baseline
    """
    out = df.sort_values(
        ["request_id", "api_model_score", "property_id"],
        ascending=[True, False, True],
        kind="mergesort",
    ).copy()

    out["baseline_rank"] = out.groupby("request_id").cumcount() + 1
    return out[out["baseline_rank"] <= k].copy()


def pg_topk_by_final_score(df: pd.DataFrame, k: int) -> pd.DataFrame:
    """
    PG (propertygroup) = PROVIDED:
      - final_score to dostarczony finalny biznesowy score/ranking
      - NIE wolno liczyć business_score() dla PG
    """
    out = df.sort_values(
        ["request_id", "final_score", "property_id"],
        ascending=[True, False, True],
        kind="mergesort",
    ).copy()

    out["pg_rank"] = out.groupby("request_id").cumcount() + 1
    out["pg_score"] = pd.to_numeric(out["final_score"], errors="coerce")

    return out[out["pg_rank"] <= k].copy()


def business_topk_by(
    df: pd.DataFrame,
    params: RerankParams,
    k: int,
    p_h: float,
    q_col: str,
    rank_col: str,
    score_col: str,
    desc: str,
) -> pd.DataFrame:
    """
    KUREKJ reranking: greedy_rerank + business_score.
    Additionally for each selected record we compute business_score_breakdown and
    append all fields as xai_* columns.
    """
    parts: list[pd.DataFrame] = []
    grouped = df.groupby("request_id", sort=False)

    for _, g in tqdm(grouped, total=grouped.ngroups, desc=desc, unit="request"):
        candidates: list[Candidate] = []
        for row in g.itertuples(index=False):
            candidates.append(
                Candidate(
                    property_id=int(row.property_id),
                    q=float(getattr(row, q_col)),
                    r=float(row.r),
                    g=float(row.g),
                    m=float(row.m),
                    v=int(row.v),
                    contract_type=row.contract_type,
                    cap_ratio=float(row.cap_ratio),
                    inv_id=int(row.inv_id) if pd.notna(row.inv_id) else None,
                    extra={"row_id": int(row.row_id)},
                )
            )

        ranked = greedy_rerank(candidates, params, k=k, p_h=p_h)

        selected: list[Candidate] = []
        rows: list[dict] = []

        for rank, cand in enumerate(ranked, start=1):
            # breakdown before appending to selected (to keep penalties consistent)
            bd = _call_breakdown(cand, params, p_h=p_h, selected=selected)
            bd_dict = _breakdown_to_dict(bd)

            # score: prefer breakdown's final_score; fallback to business_score
            score = bd_dict.get("final_score", None)
            if score is None:
                score = business_score(cand, params, p_h=p_h, selected=selected)

            bd_dict["final_score"] = float(score)

            # update history after scoring
            selected.append(cand)

            # avoid duplicate property_id in df (already in g)
            bd_dict.pop("property_id", None)

            # prefix xai_
            xai_cols = {f"xai_{k}": v for k, v in bd_dict.items()}

            rows.append(
                {
                    "row_id": cand.extra["row_id"],
                    rank_col: rank,
                    score_col: float(score),
                    **xai_cols,
                }
            )

        out = g.merge(pd.DataFrame(rows), on="row_id", how="inner")
        parts.append(out.sort_values(rank_col, kind="mergesort"))

    return pd.concat(parts, ignore_index=True)


# =========================
# SUMMARY
# =========================


def summary_wide_pair(
    left: pd.DataFrame,
    right: pd.DataFrame,
    k: int,
    left_rank: str,
    right_rank: str,
    left_prefix: str,
    right_prefix: str,
    left_score_col: str,
    right_score_col: str,
) -> pd.DataFrame:
    meta = (
        left[["request_id", "uuid", "create_date"]]
        .drop_duplicates("request_id")
        .set_index("request_id")
    )

    l_pid = left.pivot(index="request_id", columns=left_rank, values="property_id")
    l_pid.columns = [f"{left_prefix}_pid_{c}" for c in l_pid.columns]

    l_score = left.pivot(index="request_id", columns=left_rank, values=left_score_col)
    l_score.columns = [f"{left_prefix}_score_{c}" for c in l_score.columns]

    r_pid = right.pivot(index="request_id", columns=right_rank, values="property_id")
    r_pid.columns = [f"{right_prefix}_pid_{c}" for c in r_pid.columns]

    r_score = right.pivot(index="request_id", columns=right_rank, values=right_score_col)
    r_score.columns = [f"{right_prefix}_score_{c}" for c in r_score.columns]

    out = meta.join(l_pid).join(l_score).join(r_pid).join(r_score).reset_index()

    for i in range(1, k + 1):
        for col in (
            f"{left_prefix}_pid_{i}",
            f"{left_prefix}_score_{i}",
            f"{right_prefix}_pid_{i}",
            f"{right_prefix}_score_{i}",
        ):
            if col not in out.columns:
                out[col] = pd.NA

    def overlap(row):
        a = {row[f"{left_prefix}_pid_{i}"] for i in range(1, k + 1)}
        b = {row[f"{right_prefix}_pid_{i}"] for i in range(1, k + 1)}
        return len({x for x in a if pd.notna(x)} & {x for x in b if pd.notna(x)})

    out[f"top{k}_overlap_cnt"] = out.apply(overlap, axis=1)
    return out


# =========================
# METRICS
# =========================


def compute_pair_metrics(
    summary_df: pd.DataFrame,
    k: int,
    left_prefix: str,
    right_prefix: str,
) -> dict:
    n = len(summary_df)

    overlap = summary_df[f"top{k}_overlap_cnt"].fillna(0)

    top1_match = (
        summary_df[f"{left_prefix}_pid_1"]
        == summary_df[f"{right_prefix}_pid_1"]
    ).fillna(False)

    full_match = overlap == k

    return {
        "requests": int(n),
        "avg_overlap_ratio": float((overlap / k).mean()) if n else 0.0,
        "pct_full_topk_match": float(full_match.mean()) if n else 0.0,
        "pct_top1_match": float(top1_match.mean()) if n else 0.0,
    }


def build_metrics_sheet(
    k: int,
    sum_base_pg: pd.DataFrame,
    sum_base_kurekj: pd.DataFrame,
    sum_pg_kurekj: pd.DataFrame,
) -> pd.DataFrame:
    rows = []

    rows.append({
        "comparison": "baseline vs propertygroup",
        **compute_pair_metrics(sum_base_pg, k, "base", "pg"),
    })
    rows.append({
        "comparison": "baseline vs kurekj",
        **compute_pair_metrics(sum_base_kurekj, k, "base", "kurekj"),
    })
    rows.append({
        "comparison": "propertygroup vs kurekj",
        **compute_pair_metrics(sum_pg_kurekj, k, "pg", "kurekj"),
    })

    df = pd.DataFrame(rows)

    df["avg_overlap_ratio_pct"] = (df["avg_overlap_ratio"] * 100).round(2)
    df["pct_full_topk_match_pct"] = (df["pct_full_topk_match"] * 100).round(2)
    df["pct_top1_match_pct"] = (df["pct_top1_match"] * 100).round(2)

    ordered = [
        "comparison",
        "requests",
        "avg_overlap_ratio",
        "avg_overlap_ratio_pct",
        "pct_full_topk_match",
        "pct_full_topk_match_pct",
        "pct_top1_match",
        "pct_top1_match_pct",
    ]

    return df[ordered]


# =========================
# PROFIT / EARNINGS PROXY
# =========================


def _series_or_nan(df: pd.DataFrame, col: str) -> pd.Series:
    """Return df[col] if exists else a NaN series (same length)."""
    if col in df.columns:
        return df[col]
    return pd.Series(np.nan, index=df.index)


def get_position_weights(k: int) -> Dict[int, float]:
    """Position discount weights for TOP-K aggregation."""
    if POSITION_WEIGHT_MODE == "equal":
        return {i: 1.0 for i in range(1, k + 1)}

    if POSITION_WEIGHT_MODE == "custom":
        return {i: float(CUSTOM_POSITION_WEIGHTS.get(i, 0.0)) for i in range(1, k + 1)}

    return {i: float(1.0 / np.log2(i + 1)) for i in range(1, k + 1)}


def compute_effective_lead_price(df: pd.DataFrame) -> pd.Series:
    """
    One unified 'price per lead' to compare earnings across settlement models:
      - CPL: lead_price
      - Subscription/flat: estimated_lead_price (fallbacks if missing)
    """
    lead_price = pd.to_numeric(_series_or_nan(df, "lead_price"), errors="coerce")
    est_lead_price = pd.to_numeric(_series_or_nan(df, "estimated_lead_price"), errors="coerce")

    settlement = pd.to_numeric(_series_or_nan(df, "settlement_type"), errors="coerce")
    contract = _series_or_nan(df, "contract_type").astype(str).str.lower()

    if settlement.notna().any():
        is_cpl = settlement.eq(0)
    else:
        is_cpl = contract.eq(CONTRACT_TYPE_CPL)

    eff = lead_price.where(is_cpl, est_lead_price)

    lead_limit = pd.to_numeric(_series_or_nan(df, "lead_limit"), errors="coerce").replace({0: np.nan})
    est_budget = pd.to_numeric(_series_or_nan(df, "estimated_monthly_budget"), errors="coerce")
    monthly_budget = pd.to_numeric(_series_or_nan(df, "monthly_budget"), errors="coerce")
    leads_est = pd.to_numeric(_series_or_nan(df, "leads_estimation"), errors="coerce").replace({0: np.nan})

    budget_any = est_budget.fillna(monthly_budget)
    fallback1 = budget_any / lead_limit
    fallback2 = budget_any / leads_est

    eff = eff.fillna(fallback1).fillna(fallback2)

    return eff.astype(float)


def prepare_profit_frame(
    topk_df: pd.DataFrame,
    *,
    algo: str,
    rank_col: str,
    k: int,
    prob_col: str = PROFIT_PROB_COL,
) -> pd.DataFrame:
    """
    Add monetization columns to a TOP-K dataframe for aggregated 'earnings' comparison.

    Expected revenue proxy per row:
      effective_lead_price * prob(q) * position_weight * billable_flag
    """
    out = topk_df.copy()

    pos_w = get_position_weights(k)
    out["algo"] = algo
    out["rank"] = pd.to_numeric(out[rank_col], errors="coerce").astype("Int64")

    out["effective_lead_price"] = compute_effective_lead_price(out)

    out["position_weight"] = (
        out["rank"].astype(float).map(pos_w).fillna(0.0).astype(float)
    )

    q = pd.to_numeric(_series_or_nan(out, prob_col), errors="coerce").fillna(0.0)
    out["q_prob"] = q.clip(0.0, 1.0)

    cap_ratio = pd.to_numeric(_series_or_nan(out, "cap_ratio"), errors="coerce").fillna(0.0)
    out["is_billable"] = (cap_ratio < BILLABLE_CAP_THRESHOLD).astype(int)

    price = out["effective_lead_price"].fillna(0.0).clip(lower=0.0)
    out["unit_value"] = price
    out["unit_value_pos"] = price * out["position_weight"] * out["is_billable"]
    out["expected_revenue"] = price * out["q_prob"] * out["position_weight"] * out["is_billable"]

    return out


def profit_summary_row(df: pd.DataFrame) -> Dict[str, object]:
    reqs = int(df["request_id"].nunique()) if "request_id" in df.columns else 0
    rows = int(len(df))

    eff_price = pd.to_numeric(_series_or_nan(df, "effective_lead_price"), errors="coerce")
    missing_price_pct = float(eff_price.isna().mean()) if rows else 0.0

    out = {
        "algo": str(df["algo"].iloc[0]) if rows and "algo" in df.columns else "",
        "requests": reqs,
        "rows_topk": rows,
        "avg_rows_per_request": float(rows / reqs) if reqs else 0.0,
        "sum_unit_value_pos": float(
            pd.to_numeric(_series_or_nan(df, "unit_value_pos"), errors="coerce").fillna(0.0).sum()),
        "sum_expected_revenue": float(
            pd.to_numeric(_series_or_nan(df, "expected_revenue"), errors="coerce").fillna(0.0).sum()),
        "avg_expected_revenue_per_request": float(
            pd.to_numeric(_series_or_nan(df, "expected_revenue"), errors="coerce").fillna(0.0).sum() / reqs
        ) if reqs else 0.0,
        "pct_missing_effective_lead_price": round(missing_price_pct * 100, 2),
        "pct_billable_rows": round(
            float(pd.to_numeric(_series_or_nan(df, "is_billable"), errors="coerce").fillna(0).mean()) * 100,
            2) if rows else 0.0,
        "pct_cpl_rows": round(
            float((_series_or_nan(df, "contract_type").astype(str).str.lower() == CONTRACT_TYPE_CPL).mean()) * 100,
            2) if rows else 0.0,
        "pct_flat_rows": round(
            float((_series_or_nan(df, "contract_type").astype(str).str.lower() == CONTRACT_TYPE_FLAT).mean()) * 100,
            2) if rows else 0.0,
    }
    return out


def build_profit_tables(
    *,
    baseline_top: pd.DataFrame,
    pg_top: pd.DataFrame,
    kurekj_top: pd.DataFrame,
    k: int,
) -> Dict[str, pd.DataFrame]:
    """Build all profit comparison tables for the Excel report."""
    base_p = prepare_profit_frame(baseline_top, algo="baseline", rank_col="baseline_rank", k=k)
    pg_p = prepare_profit_frame(pg_top, algo="pg", rank_col="pg_rank", k=k)
    kurekj_p = prepare_profit_frame(kurekj_top, algo="kurekj", rank_col="kurekj_rank", k=k)

    overall = pd.DataFrame(
        [
            profit_summary_row(base_p),
            profit_summary_row(pg_p),
            profit_summary_row(kurekj_p),
        ]
    )

    pg_req = pg_p.groupby("request_id", as_index=True)["expected_revenue"].sum()
    k_req = kurekj_p.groupby("request_id", as_index=True)["expected_revenue"].sum()

    pair = (
        pd.DataFrame({"pg_expected_revenue": pg_req, "kurekj_expected_revenue": k_req})
        .fillna(0.0)
        .reset_index()
    )
    pair["diff_pg_minus_kurekj"] = pair["pg_expected_revenue"] - pair["kurekj_expected_revenue"]

    n = len(pair)
    pct_pg_win = float((pair["diff_pg_minus_kurekj"] > 0).mean()) if n else 0.0
    pct_k_win = float((pair["diff_pg_minus_kurekj"] < 0).mean()) if n else 0.0
    pct_tie = float((pair["diff_pg_minus_kurekj"] == 0).mean()) if n else 0.0

    pairwise = pd.DataFrame(
        [
            {"metric": "requests_compared", "value": int(n)},
            {"metric": "pct_requests_pg_higher", "value": round(pct_pg_win * 100, 2)},
            {"metric": "pct_requests_kurekj_higher", "value": round(pct_k_win * 100, 2)},
            {"metric": "pct_requests_equal", "value": round(pct_tie * 100, 2)},
            {"metric": "avg_diff_pg_minus_kurekj_per_request", "value": float(pair["diff_pg_minus_kurekj"].mean()) if n else 0.0},
            {"metric": "median_diff_pg_minus_kurekj_per_request", "value": float(pair["diff_pg_minus_kurekj"].median()) if n else 0.0},
            {"metric": "total_diff_pg_minus_kurekj", "value": float(pair["diff_pg_minus_kurekj"].sum()) if n else 0.0},
        ]
    )

    by_rank = (
        pd.concat([base_p, pg_p, kurekj_p], ignore_index=True)
        .groupby(["algo", "rank"], as_index=False)
        .agg(
            rows=("request_id", "count"),
            sum_unit_value_pos=("unit_value_pos", "sum"),
            sum_expected_revenue=("expected_revenue", "sum"),
            avg_expected_revenue=("expected_revenue", "mean"),
        )
        .sort_values(["algo", "rank"], kind="mergesort")
    )

    by_contract_src = pd.concat([base_p, pg_p, kurekj_p], ignore_index=True).copy()
    if "contract_type" not in by_contract_src.columns:
        by_contract_src["contract_type"] = ""

    by_contract_src["contract_type"] = by_contract_src["contract_type"].astype(str)

    by_contract = (
        by_contract_src
        .groupby(["algo", "contract_type"], as_index=False)
        .agg(
            rows=("request_id", "count"),
            sum_unit_value_pos=("unit_value_pos", "sum"),
            sum_expected_revenue=("expected_revenue", "sum"),
            avg_expected_revenue=("expected_revenue", "mean"),
        )
        .sort_values(["algo", "contract_type"], kind="mergesort")
    )

    pg_inv = pg_p.groupby("inv_id", as_index=True)["expected_revenue"].sum()
    k_inv = kurekj_p.groupby("inv_id", as_index=True)["expected_revenue"].sum()

    inv = (
        pd.DataFrame({"pg_expected_revenue": pg_inv, "kurekj_expected_revenue": k_inv})
        .fillna(0.0)
    )
    inv["diff_pg_minus_kurekj"] = inv["pg_expected_revenue"] - inv["kurekj_expected_revenue"]
    inv["abs_diff"] = inv["diff_pg_minus_kurekj"].abs()

    inv_top = (
        inv.sort_values("abs_diff", ascending=False, kind="mergesort")
        .head(50)
        .reset_index()
        .rename(columns={"inv_id": "billinggroup_or_offer_id"})
        .drop(columns=["abs_diff"])
    )

    return {
        "overall": overall,
        "pairwise": pairwise,
        "by_rank": by_rank,
        "by_contract": by_contract,
        "by_billinggroup_topdiff": inv_top,
    }


def write_df_section(
    xw: pd.ExcelWriter,
    sheet_name: str,
    title: str,
    df: pd.DataFrame,
    *,
    startrow: int,
) -> int:
    """Write a titled dataframe section into a single sheet; returns next startrow."""
    df.to_excel(xw, sheet_name=sheet_name, index=False, startrow=startrow + 1)

    book = xw.book
    ws = xw.sheets[sheet_name]

    is_xlsxwriter = hasattr(book, "add_format") and hasattr(ws, "write")
    if is_xlsxwriter:
        title_fmt = book.add_format({"bold": True, "font_size": 12})
        ws.write(startrow, 0, title, title_fmt)
    else:
        from openpyxl.styles import Font
        cell = ws.cell(row=startrow + 1, column=1)
        cell.value = title
        cell.font = Font(bold=True, size=12)

    return startrow + len(df) + 3


def write_profit_sheet(
    xw: pd.ExcelWriter,
    *,
    sheet_name: str,
    tables: Dict[str, pd.DataFrame],
) -> None:
    """Create one sheet with multiple profit-comparison sections."""
    start = 0
    start = write_df_section(
        xw,
        sheet_name,
        "Overall (expected revenue proxy, using effective_lead_price + q_prob + position_weight)",
        tables["overall"],
        startrow=start,
    )
    start = write_df_section(
        xw,
        sheet_name,
        "PG vs KUREKJ (per-request win rate on expected revenue)",
        tables["pairwise"],
        startrow=start,
    )
    start = write_df_section(
        xw,
        sheet_name,
        "Breakdown by rank (position)",
        tables["by_rank"],
        startrow=start,
    )
    start = write_df_section(
        xw,
        sheet_name,
        "Breakdown by contract_type (cpl vs flat)",
        tables["by_contract"],
        startrow=start,
    )
    _ = write_df_section(
        xw,
        sheet_name,
        "Top billing groups (inv_id) by |diff| in expected revenue (PG - KUREKJ)",
        tables["by_billinggroup_topdiff"],
        startrow=start,
    )


# =========================
# TRADEOFF: SCORE vs REVENUE
# =========================


def _safe_qcut_bins(x: pd.Series, q: int) -> pd.Series:
    """
    Robust q-cut bins (1..q) with tie-breaking via rank(method='first').
    Returns Int64 series.
    """
    s = pd.to_numeric(x, errors="coerce")
    if int(s.notna().sum()) == 0:
        return pd.Series(1, index=s.index, dtype="Int64")

    r = s.rank(method="first")
    unique = int(r.nunique(dropna=True))
    if unique < 2:
        return pd.Series(1, index=s.index, dtype="Int64")

    q_eff = int(min(max(int(q), 1), unique))
    if q_eff < 2:
        return pd.Series(1, index=s.index, dtype="Int64")

    try:
        b = pd.qcut(r, q=q_eff, labels=False, duplicates="drop")
    except Exception:
        b = pd.cut(r, bins=q_eff, labels=False)

    out = pd.to_numeric(b, errors="coerce").fillna(0).astype(int) + 1
    return out.astype("Int64")


def build_tradeoff_request_level(
    *,
    base_p: pd.DataFrame,
    pg_p: pd.DataFrame,
    kurekj_p: pd.DataFrame,
    k: int,
) -> pd.DataFrame:
    """
    Per-request (algo x request_id):
      - expected_revenue (sum TOP-K)
      - quality/score: mean/median + wmean (position_weight)
      - dodatkowo: pct_billable, avg_price, avg_cap_ratio, rank_score mean/wmean

    + USER_FIT (proxy ryzyka reklamacji):
      - orig_rank_*: z którego miejsca w baseline (api_model_score) pochodzi TOP-1 / TOP-K
      - user_fit_*: 0..1, im wyżej tym bliżej baseline (mniej "agresywnego" biznesowego przesunięcia)
    """
    _ = k  # future-proof

    df = pd.concat([base_p, pg_p, kurekj_p], ignore_index=True).copy()

    # ensure rank exists (prepared in prepare_profit_frame)
    df["rank"] = pd.to_numeric(_series_or_nan(df, "rank"), errors="coerce").astype("Int64")

    df["expected_revenue"] = (
        pd.to_numeric(_series_or_nan(df, "expected_revenue"), errors="coerce")
        .fillna(0.0)
        .astype(float)
    )
    df["unit_value_pos"] = (
        pd.to_numeric(_series_or_nan(df, "unit_value_pos"), errors="coerce")
        .fillna(0.0)
        .astype(float)
    )
    df["position_weight"] = (
        pd.to_numeric(_series_or_nan(df, "position_weight"), errors="coerce")
        .fillna(0.0)
        .astype(float)
    )
    df["is_billable"] = (
        pd.to_numeric(_series_or_nan(df, "is_billable"), errors="coerce")
        .fillna(0)
        .astype(int)
    )

    df["q_prob"] = (
        pd.to_numeric(_series_or_nan(df, "q_prob"), errors="coerce")
        .fillna(0.0)
        .clip(0.0, 1.0)
        .astype(float)
    )
    df["q_api"] = (
        pd.to_numeric(_series_or_nan(df, "q_api"), errors="coerce")
        .fillna(0.0)
        .clip(-1.0, 1.0)
        .astype(float)
    )
    df["q_final"] = (
        pd.to_numeric(_series_or_nan(df, "q_final"), errors="coerce")
        .fillna(0.0)
        .clip(-1.0, 1.0)
        .astype(float)
    )

    df["effective_lead_price"] = pd.to_numeric(_series_or_nan(df, "effective_lead_price"), errors="coerce").astype(float)
    df["cap_ratio"] = pd.to_numeric(_series_or_nan(df, "cap_ratio"), errors="coerce").astype(float)

    # -------------------------
    # USER_FIT columns (baseline rank distance)
    # -------------------------
    # baseline_rank_full + n_candidates should come from recs via add_baseline_full_ranks().
    # Fallback: keep NaNs (user_fit will become 0 in those cases - conservative).
    br = pd.to_numeric(_series_or_nan(df, "baseline_rank_full"), errors="coerce").astype(float)
    nc = pd.to_numeric(_series_or_nan(df, "n_candidates"), errors="coerce").astype(float)

    # conservative fill: missing baseline rank -> treat as worst rank (n_candidates)
    orig_rank = br.fillna(nc)

    # if still missing -> mark as worst later
    denom = np.where(nc <= 1.0, 1.0, nc - 1.0)  # avoid division by 0 for single-candidate requests
    orig_rank_pct = (orig_rank - 1.0) / denom

    # missing/inf -> worst (pct=1 => user_fit=0)
    orig_rank_pct = orig_rank_pct.where(np.isfinite(orig_rank_pct), 1.0).clip(0.0, 1.0)
    orig_rank_pct = orig_rank_pct.where(orig_rank.notna(), 1.0)

    user_fit = (1.0 - orig_rank_pct).clip(0.0, 1.0)

    df["baseline_rank_full"] = br.astype("Int64")
    df["n_candidates"] = nc.astype("Int64")

    df["orig_rank"] = orig_rank
    df["orig_rank_pct"] = orig_rank_pct
    df["user_fit"] = user_fit

    df["rank_displacement"] = df["orig_rank"] - pd.to_numeric(df["rank"], errors="coerce").astype(float)

    df["orig_rank_w"] = df["orig_rank"] * df["position_weight"]
    df["orig_rank_pct_w"] = df["orig_rank_pct"] * df["position_weight"]
    df["user_fit_w"] = df["user_fit"] * df["position_weight"]
    df["rank_displacement_w"] = df["rank_displacement"] * df["position_weight"]

    # -------------------------
    # rank_score unify (różne algorytmy mają różne kolumny)
    # -------------------------
    api_score = pd.to_numeric(_series_or_nan(df, "api_model_score"), errors="coerce")
    pg_score = pd.to_numeric(_series_or_nan(df, "pg_score"), errors="coerce")
    final_score = pd.to_numeric(_series_or_nan(df, "final_score"), errors="coerce")
    pg_rank_score = pg_score.fillna(final_score)
    k_score = pd.to_numeric(_series_or_nan(df, "kurekj_score"), errors="coerce")

    rank_score = pd.Series(np.nan, index=df.index, dtype=float)
    mask = df["algo"].astype(str) == "baseline"
    rank_score.loc[mask] = api_score.loc[mask]
    mask = df["algo"].astype(str) == "pg"
    rank_score.loc[mask] = pg_rank_score.loc[mask]
    mask = df["algo"].astype(str) == "kurekj"
    rank_score.loc[mask] = k_score.loc[mask]
    df["rank_score"] = pd.to_numeric(rank_score, errors="coerce").astype(float)

    df["q_prob_w"] = df["q_prob"] * df["position_weight"]
    df["q_api_w"] = df["q_api"] * df["position_weight"]
    df["q_final_w"] = df["q_final"] * df["position_weight"]
    df["rank_score_w"] = df["rank_score"] * df["position_weight"]

    req = (
        df.groupby(["algo", "request_id"], as_index=False)
        .agg(
            rows=("request_id", "count"),
            expected_revenue=("expected_revenue", "sum"),
            unit_value_pos=("unit_value_pos", "sum"),
            pct_billable=("is_billable", "mean"),
            avg_effective_lead_price=("effective_lead_price", "mean"),
            avg_cap_ratio=("cap_ratio", "mean"),
            q_prob_mean=("q_prob", "mean"),
            q_prob_median=("q_prob", "median"),
            q_api_mean=("q_api", "mean"),
            q_api_median=("q_api", "median"),
            q_final_mean=("q_final", "mean"),
            q_final_median=("q_final", "median"),
            rank_score_mean=("rank_score", "mean"),
            rank_score_median=("rank_score", "median"),
            n_candidates=("n_candidates", "max"),
            orig_rank_mean=("orig_rank", "mean"),
            orig_rank_median=("orig_rank", "median"),
            orig_rank_max=("orig_rank", "max"),
            orig_rank_pct_mean=("orig_rank_pct", "mean"),
            user_fit_mean=("user_fit", "mean"),
            user_fit_median=("user_fit", "median"),
            user_fit_min=("user_fit", "min"),
            rank_displacement_mean=("rank_displacement", "mean"),
            rank_displacement_max=("rank_displacement", "max"),
            pos_w_sum=("position_weight", "sum"),
            q_prob_w_sum=("q_prob_w", "sum"),
            q_api_w_sum=("q_api_w", "sum"),
            q_final_w_sum=("q_final_w", "sum"),
            rank_score_w_sum=("rank_score_w", "sum"),
            orig_rank_w_sum=("orig_rank_w", "sum"),
            orig_rank_pct_w_sum=("orig_rank_pct_w", "sum"),
            user_fit_w_sum=("user_fit_w", "sum"),
            rank_displacement_w_sum=("rank_displacement_w", "sum"),
        )
    )

    denom_w = pd.to_numeric(req["pos_w_sum"], errors="coerce").replace({0.0: np.nan})
    req["q_prob_wmean"] = (req["q_prob_w_sum"] / denom_w).fillna(0.0).astype(float)
    req["q_api_wmean"] = (req["q_api_w_sum"] / denom_w).fillna(0.0).astype(float)
    req["q_final_wmean"] = (req["q_final_w_sum"] / denom_w).fillna(0.0).astype(float)
    req["rank_score_wmean"] = (req["rank_score_w_sum"] / denom_w).astype(float)

    # USER_FIT weighted means
    req["orig_rank_wmean"] = (req["orig_rank_w_sum"] / denom_w).astype(float)
    req["orig_rank_pct_wmean"] = (req["orig_rank_pct_w_sum"] / denom_w).fillna(1.0).clip(0.0, 1.0).astype(float)
    req["user_fit_wmean"] = (req["user_fit_w_sum"] / denom_w).fillna(0.0).clip(0.0, 1.0).astype(float)
    req["rank_displacement_wmean"] = (req["rank_displacement_w_sum"] / denom_w).astype(float)

    # TOP-1 specifics (rank==1)
    top1 = (
        df.loc[df["rank"] == 1, ["algo", "request_id", "orig_rank", "orig_rank_pct", "user_fit", "rank_displacement"]]
        .rename(
            columns={
                "orig_rank": "orig_rank_top1",
                "orig_rank_pct": "orig_rank_pct_top1",
                "user_fit": "user_fit_top1",
                "rank_displacement": "rank_displacement_top1",
            }
        )
        .copy()
    )

    req = req.merge(top1, on=["algo", "request_id"], how="left")

    req["orig_rank_top1"] = pd.to_numeric(req["orig_rank_top1"], errors="coerce").astype("Int64")
    req["is_top1_from_baseline_safe"] = (
        pd.to_numeric(req["orig_rank_top1"], errors="coerce").le(USER_FIT_SAFE_TOPN).fillna(False).astype(int)
    )

    # how many positions TOP-1 was moved up vs baseline (positive = moved up)
    req["delta_top1_vs_baseline"] = (
        pd.to_numeric(req["orig_rank_top1"], errors="coerce").astype(float) - 1.0
    )

    req["avg_expected_revenue_per_row"] = (
        pd.to_numeric(req["expected_revenue"], errors="coerce").fillna(0.0)
        / pd.to_numeric(req["rows"], errors="coerce").replace({0: np.nan})
    ).fillna(0.0).astype(float)

    return req.sort_values(["algo", "expected_revenue"], ascending=[True, False], kind="mergesort").reset_index(drop=True)

def build_tradeoff_bins(
    req_df: pd.DataFrame,
    *,
    bins: int,
    mode: str = "per_algo",
    revenue_col: str = "expected_revenue",
) -> pd.DataFrame:
    if mode not in {"per_algo", "global"}:
        raise ValueError(f"TRADEOFF_BIN_MODE must be 'per_algo' or 'global', got: {mode!r}")

    df = req_df.copy()
    if mode == "global":
        df["revenue_bin"] = _safe_qcut_bins(df[revenue_col], bins)
    else:
        parts = []
        for algo, g in df.groupby("algo", sort=True):
            t = g.copy()
            t["revenue_bin"] = _safe_qcut_bins(t[revenue_col], bins)
            parts.append(t)
        df = pd.concat(parts, ignore_index=True) if parts else df.assign(revenue_bin=pd.NA)

    out = (
        df.groupby(["algo", "revenue_bin"], as_index=False)
        .agg(
            requests=("request_id", "count"),
            expected_revenue_mean=(revenue_col, "mean"),
            expected_revenue_median=(revenue_col, "median"),
            expected_revenue_min=(revenue_col, "min"),
            expected_revenue_max=(revenue_col, "max"),
            q_prob_wmean_mean=("q_prob_wmean", "mean"),
            q_prob_wmean_median=("q_prob_wmean", "median"),
            q_api_wmean_mean=("q_api_wmean", "mean"),
            q_api_wmean_median=("q_api_wmean", "median"),
            rank_score_wmean_mean=("rank_score_wmean", "mean"),
            rank_score_wmean_median=("rank_score_wmean", "median"),
            pct_billable_mean=("pct_billable", "mean"),
            avg_effective_lead_price_mean=("avg_effective_lead_price", "mean"),
            avg_cap_ratio_mean=("avg_cap_ratio", "mean"),

            # USER_FIT (if present in req_df)
            user_fit_wmean_mean=("user_fit_wmean", "mean"),
            user_fit_wmean_median=("user_fit_wmean", "median"),
            user_fit_top1_mean=("user_fit_top1", "mean"),
            user_fit_top1_median=("user_fit_top1", "median"),
            orig_rank_top1_mean=("orig_rank_top1", "mean"),
            orig_rank_top1_median=("orig_rank_top1", "median"),
            is_top1_from_baseline_safe_mean=("is_top1_from_baseline_safe", "mean"),
        )
        .sort_values(["algo", "revenue_bin"], kind="mergesort")
        .reset_index(drop=True)
    )

    for c in [
        "expected_revenue_mean",
        "expected_revenue_median",
        "expected_revenue_min",
        "expected_revenue_max",
        "avg_effective_lead_price_mean",
        "avg_cap_ratio_mean",
        "pct_billable_mean",
        "user_fit_wmean_mean",
        "user_fit_wmean_median",
        "user_fit_top1_mean",
        "user_fit_top1_median",
        "orig_rank_top1_mean",
        "orig_rank_top1_median",
        "is_top1_from_baseline_safe_mean",
    ]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").astype(float)

    return out


def build_tradeoff_quadrants_vs_baseline(
    req_df: pd.DataFrame,
    *,
    baseline_algo: str,
    compare_algos: List[str],
    revenue_col: str,
    quality_col: str,
    eps: float,
) -> pd.DataFrame:
    """
    „Koszt” wprost: ile % requestów ma revenue_up ale quality_down vs baseline?
    """
    wide_rev = req_df.pivot(index="request_id", columns="algo", values=revenue_col).fillna(0.0)
    wide_q = req_df.pivot(index="request_id", columns="algo", values=quality_col).fillna(0.0)

    rows = []
    for algo in compare_algos:
        if baseline_algo not in wide_rev.columns or algo not in wide_rev.columns:
            continue
        if baseline_algo not in wide_q.columns or algo not in wide_q.columns:
            continue

        dr = pd.to_numeric(wide_rev[algo], errors="coerce").fillna(0.0) - pd.to_numeric(wide_rev[baseline_algo], errors="coerce").fillna(0.0)
        dq = pd.to_numeric(wide_q[algo], errors="coerce").fillna(0.0) - pd.to_numeric(wide_q[baseline_algo], errors="coerce").fillna(0.0)

        n = int(len(dr))
        if n == 0:
            continue

        rev_up = dr > eps
        rev_down = dr < -eps
        rev_tie = ~(rev_up | rev_down)

        q_up = dq > eps
        q_down = dq < -eps
        q_tie = ~(q_up | q_down)

        rows.append(
            {
                "compare_algo": algo,
                "baseline_algo": baseline_algo,
                "quality_col": quality_col,
                "requests_compared": n,
                "revenue_up_pct": float(rev_up.mean() * 100),
                "revenue_down_pct": float(rev_down.mean() * 100),
                "revenue_tie_pct": float(rev_tie.mean() * 100),
                "quality_up_pct": float(q_up.mean() * 100),
                "quality_down_pct": float(q_down.mean() * 100),
                "quality_tie_pct": float(q_tie.mean() * 100),
                "rev_up_quality_up_pct": float((rev_up & q_up).mean() * 100),
                "rev_up_quality_down_pct": float((rev_up & q_down).mean() * 100),
                "rev_down_quality_up_pct": float((rev_down & q_up).mean() * 100),
                "rev_down_quality_down_pct": float((rev_down & q_down).mean() * 100),
                "mean_delta_revenue": float(dr.mean()),
                "median_delta_revenue": float(dr.median()),
                "mean_delta_quality": float(dq.mean()),
                "median_delta_quality": float(dq.median()),
                "baseline_revenue_mean": float(pd.to_numeric(wide_rev[baseline_algo], errors="coerce").fillna(0.0).mean()),
                "algo_revenue_mean": float(pd.to_numeric(wide_rev[algo], errors="coerce").fillna(0.0).mean()),
                "baseline_quality_mean": float(pd.to_numeric(wide_q[baseline_algo], errors="coerce").fillna(0.0).mean()),
                "algo_quality_mean": float(pd.to_numeric(wide_q[algo], errors="coerce").fillna(0.0).mean()),
            }
        )

    return pd.DataFrame(rows).sort_values(["compare_algo"], kind="mergesort").reset_index(drop=True)


def build_tradeoff_correlations(
    req_df: pd.DataFrame,
    *,
    revenue_col: str,
    quality_cols: List[str],
) -> pd.DataFrame:
    """
    Korelacje revenue vs score (Pearson + Spearman) per algo.
    """
    rows = []
    for algo, g in req_df.groupby("algo", sort=True):
        for qc in quality_cols:
            if qc not in g.columns or revenue_col not in g.columns:
                continue
            x = pd.to_numeric(g[revenue_col], errors="coerce")
            y = pd.to_numeric(g[qc], errors="coerce")
            m = x.notna() & y.notna()
            n = int(m.sum())
            if n < 2:
                pearson = np.nan
                spearman = np.nan
            else:
                tmp = pd.DataFrame({"x": x[m], "y": y[m]})
                pearson = float(tmp.corr(method="pearson").iloc[0, 1])
                spearman = float(tmp.corr(method="spearman").iloc[0, 1])
            rows.append(
                {
                    "algo": str(algo),
                    "quality_col": str(qc),
                    "n": n,
                    "pearson": pearson,
                    "spearman": spearman,
                }
            )

    return (
        pd.DataFrame(rows)
        .sort_values(["algo", "quality_col"], kind="mergesort")
        .reset_index(drop=True)
    )




# =========================
# USER FIT TABLES (complaint-risk proxy)
# =========================


def build_user_fit_overall(
    req_df: pd.DataFrame,
    *,
    safe_topn: int = USER_FIT_SAFE_TOPN,
    thresholds: List[int] = USER_FIT_THRESHOLDS,
) -> pd.DataFrame:
    """
    Aggregated per-algorithm summary of USER_FIT.
    """
    df = req_df.copy()

    if "orig_rank_top1" not in df.columns:
        raise ValueError("build_user_fit_overall: missing column 'orig_rank_top1' (did you run build_tradeoff_request_level?)")

    rows = []
    for algo, g in df.groupby("algo", sort=True):
        r1 = pd.to_numeric(g["orig_rank_top1"], errors="coerce").astype(float)
        uf_k = pd.to_numeric(_series_or_nan(g, "user_fit_wmean"), errors="coerce").astype(float)
        uf_1 = pd.to_numeric(_series_or_nan(g, "user_fit_top1"), errors="coerce").astype(float)
        rev = pd.to_numeric(_series_or_nan(g, "expected_revenue"), errors="coerce").astype(float)

        n = int(len(g))
        n_r1 = int(r1.notna().sum())

        row: Dict[str, Any] = {
            "algo": str(algo),
            "requests": n,
            "pct_missing_orig_rank_top1": round(float((1 - (n_r1 / n)) * 100), 2) if n else 0.0,
            "expected_revenue_mean": float(rev.mean()) if n else 0.0,
            "expected_revenue_median": float(rev.median()) if n else 0.0,
            "orig_rank_top1_mean": float(r1.mean()) if n_r1 else np.nan,
            "orig_rank_top1_median": float(r1.median()) if n_r1 else np.nan,
            "orig_rank_top1_p90": float(r1.quantile(0.9)) if n_r1 else np.nan,
            "orig_rank_top1_max": float(r1.max()) if n_r1 else np.nan,
            "pct_top1_from_baseline_safe": round(float((r1 <= safe_topn).mean() * 100), 2) if n_r1 else 0.0,
            "user_fit_wmean_mean": float(uf_k.mean()) if n else 0.0,
            "user_fit_wmean_median": float(uf_k.median()) if n else 0.0,
            "user_fit_top1_mean": float(uf_1.mean()) if n else 0.0,
            "user_fit_top1_median": float(uf_1.median()) if n else 0.0,
        }

        for t in thresholds:
            row[f"pct_top1_baseline_rank_le_{t}"] = round(float((r1 <= float(t)).mean() * 100), 2) if n_r1 else 0.0
            row[f"pct_top1_baseline_rank_gt_{t}"] = round(float((r1 > float(t)).mean() * 100), 2) if n_r1 else 0.0

        rows.append(row)

    return pd.DataFrame(rows).sort_values(["algo"], kind="mergesort").reset_index(drop=True)


def build_user_fit_thresholds(
    req_df: pd.DataFrame,
    *,
    thresholds: List[int] = USER_FIT_THRESHOLDS,
) -> pd.DataFrame:
    """
    Threshold-style table: for each algo and each threshold t show pct of requests with orig_rank_top1 <= t.
    """
    df = req_df.copy()

    if "orig_rank_top1" not in df.columns:
        raise ValueError("build_user_fit_thresholds: missing column 'orig_rank_top1'")

    rows = []
    for algo, g in df.groupby("algo", sort=True):
        r1 = pd.to_numeric(g["orig_rank_top1"], errors="coerce").astype(float)
        n = int(r1.notna().sum())
        for t in thresholds:
            rows.append(
                {
                    "algo": str(algo),
                    "threshold_topn": int(t),
                    "requests_with_rank": n,
                    "pct_top1_baseline_rank_le_t": round(float((r1 <= float(t)).mean() * 100), 2) if n else 0.0,
                    "pct_top1_baseline_rank_gt_t": round(float((r1 > float(t)).mean() * 100), 2) if n else 0.0,
                }
            )

    return pd.DataFrame(rows).sort_values(["algo", "threshold_topn"], kind="mergesort").reset_index(drop=True)


def build_user_fit_bins_top1(
    req_df: pd.DataFrame,
    *,
    bins: List[float] = USER_FIT_TOP1_BINS,
    labels: List[str] = USER_FIT_TOP1_BIN_LABELS,
) -> pd.DataFrame:
    """
    Distribution of orig_rank_top1 in coarse bins (per algo).
    """
    if len(bins) != len(labels) + 1:
        raise ValueError("build_user_fit_bins_top1: bins must have len(labels)+1")

    df = req_df.copy()
    r1 = pd.to_numeric(df.get("orig_rank_top1"), errors="coerce").astype(float)

    df["orig_rank_top1_bin"] = pd.cut(
        r1,
        bins=bins,
        labels=labels,
        right=True,
        include_lowest=True,
    )

    dist = (
        df.groupby(["algo", "orig_rank_top1_bin"], dropna=False)
        .size()
        .reset_index(name="requests")
        .sort_values(["algo", "orig_rank_top1_bin"], kind="mergesort")
        .reset_index(drop=True)
    )

    dist["requests_pct"] = (
        dist.groupby("algo")["requests"]
        .transform(lambda x: x / x.sum() * 100 if float(x.sum()) > 0 else 0.0)
        .round(2)
    )

    return dist


def build_user_fit_tables(req_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Convenience wrapper."""
    return {
        "user_fit_overall": build_user_fit_overall(req_df),
        "user_fit_thresholds": build_user_fit_thresholds(req_df),
        "user_fit_bins_top1": build_user_fit_bins_top1(req_df),
    }

# =========================
# XAI TABLES (ARTICLE STATS)
# =========================


def build_xai_tables(kurekj_top: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Article-ready summaries from xai_* columns produced by business_score_breakdown.

    Outputs:
      - xai_overall: mean/mean_abs for components
      - xai_by_rank: same by kurekj_rank
      - xai_by_contract: same by contract_type
      - xai_cap_regions: counts if xai_cap_region exists
      - xai_div_reasons: counts if xai_div_reason exists
      - xai_counterfactuals: summaries for xai_diff_no_* columns if present
      - xai_recon_check: reconstruction error vs xai_final_score
    """
    df = kurekj_top.copy()

    if "xai_final_score" not in df.columns:
        raise ValueError("kurekj_top: missing xai_final_score (expected from breakdown)")

    # Component columns discovery
    component_cols: List[Tuple[str, str]] = []  # (name, impact_col)

    # core/base
    if "xai_core" in df.columns:
        df["impact_core"] = pd.to_numeric(df["xai_core"], errors="coerce").fillna(0.0).astype(float)
        component_cols.append(("core", "impact_core"))

    # contrib_*
    for c in sorted([c for c in df.columns if str(c).startswith("xai_contrib_")]):
        name = str(c).replace("xai_contrib_", "")
        impact_col = f"impact_{name}"
        df[impact_col] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype(float)
        component_cols.append((name, impact_col))

    # term_*
    for c in sorted([c for c in df.columns if str(c).startswith("xai_term_")]):
        name = str(c).replace("xai_term_", "")
        impact_col = f"impact_{name}"
        df[impact_col] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype(float)
        component_cols.append((name, impact_col))

    # Cap/div penalties (weighted preferred)
    def _add_penalty_component(name: str, col: str) -> None:
        if col not in df.columns:
            return
        s = pd.to_numeric(df[col], errors="coerce").fillna(0.0).astype(float)
        # if mostly positive -> treat as magnitude and make negative impact
        if (s >= 0).mean() > 0.90:
            impact = -s
        else:
            impact = s
        impact_col = f"impact_{name}"
        df[impact_col] = impact.astype(float)
        component_cols.append((name, impact_col))

    _add_penalty_component("cap", "xai_cap_penalty_weighted")
    _add_penalty_component("div", "xai_div_penalty_weighted")
    _add_penalty_component("cap", "xai_cap_penalty")  # fallback
    _add_penalty_component("div", "xai_div_penalty")  # fallback

    # If nothing found, fallback to "all numeric xai_* except final_score"
    if not component_cols:
        numeric_xai = []
        for c in df.columns:
            if not str(c).startswith("xai_"):
                continue
            if c in {"xai_final_score"}:
                continue
            if df[c].dtype == object:
                continue
            numeric_xai.append(c)
        for c in numeric_xai:
            name = str(c).replace("xai_", "")
            impact_col = f"impact_{name}"
            df[impact_col] = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype(float)
            component_cols.append((name, impact_col))

    if not component_cols:
        raise ValueError("No XAI components detected in kurekj_top (xai_* columns).")

    # Reconstruction check (sum of components vs final score)
    df["score_reconstructed"] = 0.0
    for _, col in component_cols:
        df["score_reconstructed"] += pd.to_numeric(df[col], errors="coerce").fillna(0.0).astype(float)
    df["recon_error"] = df["score_reconstructed"] - pd.to_numeric(df["xai_final_score"], errors="coerce").fillna(0.0)

    def summarize(frame: pd.DataFrame) -> pd.DataFrame:
        rows = []
        mean_abs_total = 0.0
        tmp: Dict[str, Dict[str, float]] = {}

        for comp, col in component_cols:
            s = pd.to_numeric(frame[col], errors="coerce").fillna(0.0).astype(float)
            d = {
                "component": comp,
                "mean": float(s.mean()),
                "mean_abs": float(s.abs().mean()),
                "median_abs": float(s.abs().median()),
                "pct_positive": float((s > 0).mean() * 100),
                "pct_negative": float((s < 0).mean() * 100),
            }
            tmp[comp] = d
            mean_abs_total += d["mean_abs"]

        for comp, d in tmp.items():
            d["share_mean_abs_pct"] = (d["mean_abs"] / mean_abs_total * 100) if mean_abs_total > 0 else 0.0
            rows.append(d)

        return (
            pd.DataFrame(rows)
            .sort_values("mean_abs", ascending=False, kind="mergesort")
            .reset_index(drop=True)
        )

    overall = summarize(df)

    if "kurekj_rank" in df.columns:
        by_rank_parts = []
        for rank, g in df.groupby("kurekj_rank", sort=True):
            t = summarize(g).copy()
            t.insert(0, "kurekj_rank", int(rank) if pd.notna(rank) else rank)
            by_rank_parts.append(t)
        by_rank = pd.concat(by_rank_parts, ignore_index=True) if by_rank_parts else pd.DataFrame()
    else:
        by_rank = pd.DataFrame()

    if "contract_type" in df.columns:
        by_contract_parts = []
        for ct, g in df.groupby("contract_type", sort=True):
            t = summarize(g).copy()
            t.insert(0, "contract_type", str(ct))
            by_contract_parts.append(t)
        by_contract = pd.concat(by_contract_parts, ignore_index=True) if by_contract_parts else pd.DataFrame()
    else:
        by_contract = pd.DataFrame()

    if "xai_cap_region" in df.columns:
        cap_regions = (
            df["xai_cap_region"].astype(str).value_counts(dropna=False)
            .rename_axis("cap_region").reset_index(name="rows")
        )
        cap_regions["rows_pct"] = (cap_regions["rows"] / len(df) * 100).round(2) if len(df) else 0.0
    else:
        cap_regions = pd.DataFrame()

    if "xai_div_reason" in df.columns:
        div_reasons = (
            df["xai_div_reason"].astype(str).value_counts(dropna=False)
            .rename_axis("div_reason").reset_index(name="rows")
        )
        div_reasons["rows_pct"] = (div_reasons["rows"] / len(df) * 100).round(2) if len(df) else 0.0
    else:
        div_reasons = pd.DataFrame()

    diff_cols = sorted([c for c in df.columns if str(c).startswith("xai_diff_no_")])
    if diff_cols:
        rows_cf = []
        for c in diff_cols:
            s = pd.to_numeric(df[c], errors="coerce").fillna(0.0).astype(float)
            rows_cf.append(
                {
                    "counterfactual": str(c).replace("xai_diff_no_", "no_"),
                    "mean": float(s.mean()),
                    "mean_abs": float(s.abs().mean()),
                    "median_abs": float(s.abs().median()),
                    "pct_positive": float((s > 0).mean() * 100),
                    "pct_negative": float((s < 0).mean() * 100),
                }
            )
        counterfactuals = (
            pd.DataFrame(rows_cf)
            .sort_values("mean_abs", ascending=False, kind="mergesort")
            .reset_index(drop=True)
        )
    else:
        counterfactuals = pd.DataFrame()

    recon_check = pd.DataFrame(
        [{
            "rows": int(len(df)),
            "components_count": int(len(component_cols)),
            "recon_error_mean": float(df["recon_error"].mean()) if len(df) else 0.0,
            "recon_error_p95_abs": float(df["recon_error"].abs().quantile(0.95)) if len(df) else 0.0,
            "recon_error_max_abs": float(df["recon_error"].abs().max()) if len(df) else 0.0,
        }]
    )

    return {
        "xai_overall": overall,
        "xai_by_rank": by_rank,
        "xai_by_contract": by_contract,
        "xai_cap_regions": cap_regions,
        "xai_div_reasons": div_reasons,
        "xai_counterfactuals": counterfactuals,
        "xai_recon_check": recon_check,
    }


# =========================
# EXCEL HELPERS (GROUP HEADER + COLORS)
# =========================


def choose_excel_engine() -> str:
    try:
        __import__("xlsxwriter")
        return "xlsxwriter"
    except Exception:
        return "openpyxl"


def add_param_columns(df: pd.DataFrame, params: RerankParams, p_h: float) -> pd.DataFrame:
    """Adds constant params columns so they can be grouped/colored in the same sheet."""
    out = df.copy()
    for k, v in asdict(params).items():
        out[f"param_{k}"] = v
    out["p_h"] = p_h
    return out


def _compress_segments(cols: List[str], col2group: Dict[str, str]) -> List[Tuple[str, int, int]]:
    """Return contiguous (group_label, start_col_idx, end_col_idx) segments."""
    if not cols:
        return []
    segs: List[Tuple[str, int, int]] = []
    cur = col2group.get(cols[0], LABEL_OTHER)
    start = 0
    for i in range(1, len(cols)):
        g = col2group.get(cols[i], LABEL_OTHER)
        if g != cur:
            segs.append((cur, start, i - 1))
            cur = g
            start = i
    segs.append((cur, start, len(cols) - 1))
    return segs


def _infer_col_groups(
    df_cols: List[str],
    *,
    recs_input_cols: set[str],
    opm_cols: set[str],
    feature_cols: set[str],
    weight_cols: set[str],
    output_cols: set[str],
) -> Dict[str, str]:
    """Map each column -> group label."""
    col2group: Dict[str, str] = {}
    for c in df_cols:
        sc = str(c)
        if sc.startswith("xai_"):
            col2group[c] = LABEL_XAI
            continue
        if sc.startswith("param_"):
            col2group[c] = LABEL_WEIGHTS
            continue

        if c in output_cols:
            col2group[c] = LABEL_OUTPUTS
        elif c in weight_cols:
            col2group[c] = LABEL_WEIGHTS
        elif c in feature_cols:
            col2group[c] = LABEL_FEATURES
        elif c in opm_cols:
            col2group[c] = LABEL_OPM
        elif c in recs_input_cols:
            col2group[c] = LABEL_RECS
        else:
            col2group[c] = LABEL_OTHER
    return col2group


def write_df_with_group_header(
    xw: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    recs_input_cols: set[str],
    opm_cols: set[str],
    feature_cols: set[str],
    weight_cols: set[str],
    output_cols: set[str],
) -> None:
    """
    Writes df to Excel with:
      - extra row 0: group labels (merged across columns)
      - row 1: normal column headers, but colored per group
      - freeze panes below row 1
    Works for xlsxwriter and openpyxl engines.
    """
    df.to_excel(xw, sheet_name=sheet_name, index=False, startrow=1)

    cols = list(df.columns)
    col2group = _infer_col_groups(
        cols,
        recs_input_cols=recs_input_cols,
        opm_cols=opm_cols,
        feature_cols=feature_cols,
        weight_cols=weight_cols,
        output_cols=output_cols,
    )
    segs = _compress_segments(cols, col2group)

    book = xw.book
    ws = xw.sheets[sheet_name]

    is_xlsxwriter = hasattr(book, "add_format") and hasattr(ws, "merge_range")

    if is_xlsxwriter:
        workbook = book

        def fmt(bg: str, *, bold: bool) -> object:
            return workbook.add_format(
                {
                    "bold": bold,
                    "align": "center",
                    "valign": "vcenter",
                    "bg_color": bg,
                    "border": 1,
                    "text_wrap": True,
                }
            )

        group_fmt = {k: fmt(GROUP_BG.get(k, GROUP_BG[LABEL_OTHER]), bold=True) for k in GROUP_BG}
        head_fmt = {k: fmt(GROUP_BG.get(k, GROUP_BG[LABEL_OTHER]), bold=True) for k in GROUP_BG}

        ws.set_row(0, 22)
        ws.set_row(1, 18)

        for g, c0, c1 in segs:
            f = group_fmt.get(g, group_fmt[LABEL_OTHER])
            if c0 == c1:
                ws.write(0, c0, g, f)
            else:
                ws.merge_range(0, c0, 0, c1, g, f)

        for j, c in enumerate(cols):
            g = col2group.get(c, LABEL_OTHER)
            f = head_fmt.get(g, head_fmt[LABEL_OTHER])
            ws.write(1, j, c, f)

        ws.freeze_panes(2, 0)

        if len(cols) > 0:
            ws.autofilter(1, 0, 1, len(cols) - 1)

    else:
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        header_row = 2
        group_row = 1

        ws.row_dimensions[group_row].height = 22
        ws.row_dimensions[header_row].height = 18

        def make_fill(hex_color: str) -> PatternFill:
            c = hex_color.replace("#", "")
            return PatternFill(patternType="solid", fgColor=c)

        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for g, c0, c1 in segs:
            bg = GROUP_BG.get(g, GROUP_BG[LABEL_OTHER])
            fill = make_fill(bg)

            start_col = c0 + 1
            end_col = c1 + 1

            if start_col != end_col:
                ws.merge_cells(
                    start_row=group_row,
                    start_column=start_col,
                    end_row=group_row,
                    end_column=end_col,
                )

            for col_i in range(start_col, end_col + 1):
                cell = ws.cell(row=group_row, column=col_i)
                cell.fill = fill
                cell.font = bold_font
                cell.alignment = center
                cell.border = border

            ws.cell(row=group_row, column=start_col).value = g

        for j, c in enumerate(cols, start=1):
            g = col2group.get(c, LABEL_OTHER)
            fill = make_fill(GROUP_BG.get(g, GROUP_BG[LABEL_OTHER]))

            cell = ws.cell(row=header_row, column=j)
            cell.value = c
            cell.fill = fill
            cell.font = bold_font
            cell.alignment = center
            cell.border = border

            col_letter = get_column_letter(j)
            ws.column_dimensions[col_letter].width = max(12, min(45, len(str(c)) + 2))

        ws.freeze_panes = ws["A3"]

        if len(cols) > 0:
            last_col_letter = get_column_letter(len(cols))
            last_row = df.shape[0] + header_row
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"


# =========================
# MAIN
# =========================


def main() -> None:
    recs = read_table(RECS_PATH)
    ensure_required_columns(recs, RECS_REQUIRED_COLS, df_name="recs")

    recs = normalize_create_date(recs)

    recs_input_cols_set = set(recs.columns)

    opm_cols_set: set[str] = set()
    if OFFERS_OPM_DATA_PATH.exists():
        offers_opm = read_offer_performance_metrics(OFFERS_OPM_DATA_PATH)
        opm_cols_set = set(offers_opm.columns) - {"offer_id"}
        recs = enrich_with_offer_performance_metrics(recs, offers_opm)
    else:
        print(
            "WARN: OFFERS_OPM_DATA_PATH does not exist -> using DEFAULT_* for business features. "
            f"path={OFFERS_OPM_DATA_PATH.resolve()}"
        )

    recs = build_features(recs)

    # Baseline ranks (needed for USER_FIT metric)
    if USER_FIT_ENABLED:
        recs = add_baseline_full_ranks(recs)

    params = RerankParams(**RERANK_PARAMS_DICT)

    baseline_top = baseline_topk_by(recs, TOPK)
    pg_top = pg_topk_by_final_score(recs, TOPK)

    kurekj_top = business_topk_by(
        recs,
        params,
        TOPK,
        P_H,
        q_col="q_api",
        rank_col="kurekj_rank",
        score_col="kurekj_score",
        desc="Kurekj reranking (business_score over api_model_score) + XAI breakdown",
    )

    baseline_top = add_param_columns(baseline_top, params, P_H)
    pg_top = add_param_columns(pg_top, params, P_H)
    kurekj_top = add_param_columns(kurekj_top, params, P_H)

    # XAI tables for article
    xai_tables = build_xai_tables(kurekj_top)

    sum_base_pg = summary_wide_pair(
        baseline_top,
        pg_top,
        TOPK,
        "baseline_rank",
        "pg_rank",
        "base",
        "pg",
        "api_model_score",
        "pg_score",
    )

    sum_base_kurekj = summary_wide_pair(
        baseline_top,
        kurekj_top,
        TOPK,
        "baseline_rank",
        "kurekj_rank",
        "base",
        "kurekj",
        "api_model_score",
        "kurekj_score",
    )

    sum_pg_kurekj = summary_wide_pair(
        pg_top,
        kurekj_top,
        TOPK,
        "pg_rank",
        "kurekj_rank",
        "pg",
        "kurekj",
        "pg_score",
        "kurekj_score",
    )

    metrics_df = build_metrics_sheet(
        TOPK,
        sum_base_pg,
        sum_base_kurekj,
        sum_pg_kurekj,
    )

    script_path = Path(__file__).resolve()
    script_filename = script_path.name
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    out_dir = script_path.parent / f"Output_[{script_filename}]_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)

    out_path = out_dir / f"{OUT_XLSX_BASENAME}_{ts}.xlsx"

    feature_cols_set = {
        "rec_day",
        "row_id",
        "q_api",
        "q_final",
        "r",
        "g",
        "m",
        "v",
        "m_budget_raw",
        "lead_limit_raw",
        "effective_lead_price",
        "contract_type",
        "cap_ratio",
        "inv_id",
        "n_candidates",
    }
    weight_cols_set = {f"param_{k}" for k in asdict(params).keys()} | {"p_h"}
    output_cols_set = {
        "baseline_rank",
        "pg_rank",
        "pg_score",
        "kurekj_rank",
        "kurekj_score",
    }

    with pd.ExcelWriter(out_path, engine=choose_excel_engine()) as xw:
        write_df_with_group_header(
            xw,
            "baseline_topk",
            baseline_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        write_df_with_group_header(
            xw,
            "pg_business_topk",
            pg_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        write_df_with_group_header(
            xw,
            "kurekj_business",
            kurekj_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        sum_base_pg.to_excel(xw, sheet_name="sum_base_vs_pg", index=False)
        sum_base_kurekj.to_excel(xw, sheet_name="sum_base_vs_kurekj", index=False)
        sum_pg_kurekj.to_excel(xw, sheet_name="sum_pg_vs_kurekj", index=False)

        metrics_df.to_excel(xw, sheet_name="metrics", index=False)

        profit_tables = build_profit_tables(
            baseline_top=baseline_top,
            pg_top=pg_top,
            kurekj_top=kurekj_top,
            k=TOPK,
        )
        write_profit_sheet(
            xw,
            sheet_name="profit_comparison",
            tables=profit_tables,
        )

        # =========================
        # Trade-off sheets (revenue vs score)
        # =========================
        base_p = prepare_profit_frame(baseline_top, algo="baseline", rank_col="baseline_rank", k=TOPK)
        pg_p = prepare_profit_frame(pg_top, algo="pg", rank_col="pg_rank", k=TOPK)
        kurekj_p = prepare_profit_frame(kurekj_top, algo="kurekj", rank_col="kurekj_rank", k=TOPK)

        tradeoff_req = build_tradeoff_request_level(base_p=base_p, pg_p=pg_p, kurekj_p=kurekj_p, k=TOPK)
        tradeoff_bins = build_tradeoff_bins(
            tradeoff_req,
            bins=TRADEOFF_BINS,
            mode=TRADEOFF_BIN_MODE,
            revenue_col="expected_revenue",
        )
        tradeoff_quadrants = build_tradeoff_quadrants_vs_baseline(
            tradeoff_req,
            baseline_algo=TRADEOFF_BASELINE_ALGO,
            compare_algos=["pg", "kurekj"],
            revenue_col="expected_revenue",
            quality_col=TRADEOFF_QUALITY_COL,
            eps=TRADEOFF_EPS,
        )
        tradeoff_quadrants_user_fit = build_tradeoff_quadrants_vs_baseline(
            tradeoff_req,
            baseline_algo=TRADEOFF_BASELINE_ALGO,
            compare_algos=["pg", "kurekj"],
            revenue_col="expected_revenue",
            quality_col=USER_FIT_QUALITY_COL,
            eps=TRADEOFF_EPS,
        )
        tradeoff_corr = build_tradeoff_correlations(
            tradeoff_req,
            revenue_col="expected_revenue",
            quality_cols=[
                TRADEOFF_QUALITY_COL,
                USER_FIT_QUALITY_COL,
                "user_fit_top1",
                "orig_rank_top1",
                "q_api_wmean",
                "rank_score_wmean",
            ],
        )

        tradeoff_req.to_excel(xw, sheet_name="tradeoff_request_level", index=False)
        tradeoff_bins.to_excel(xw, sheet_name="tradeoff_bins", index=False)
        tradeoff_quadrants.to_excel(xw, sheet_name="tradeoff_quadrants", index=False)
        tradeoff_corr.to_excel(xw, sheet_name="tradeoff_corr", index=False)

        # USER_FIT sheets
        tradeoff_quadrants_user_fit.to_excel(xw, sheet_name="tradeoff_quadrants_user_fit", index=False)

        user_fit_tables = build_user_fit_tables(tradeoff_req)
        user_fit_tables["user_fit_overall"].to_excel(xw, sheet_name="user_fit_overall", index=False)
        user_fit_tables["user_fit_thresholds"].to_excel(xw, sheet_name="user_fit_thresholds", index=False)
        user_fit_tables["user_fit_bins_top1"].to_excel(xw, sheet_name="user_fit_bins_top1", index=False)


        # XAI sheets
        xai_tables["xai_overall"].to_excel(xw, sheet_name="xai_overall", index=False)
        xai_tables["xai_by_rank"].to_excel(xw, sheet_name="xai_by_rank", index=False)
        xai_tables["xai_by_contract"].to_excel(xw, sheet_name="xai_by_contract", index=False)
        xai_tables["xai_cap_regions"].to_excel(xw, sheet_name="xai_cap_regions", index=False)
        xai_tables["xai_div_reasons"].to_excel(xw, sheet_name="xai_div_reasons", index=False)
        xai_tables["xai_counterfactuals"].to_excel(xw, sheet_name="xai_counterfactuals", index=False)
        xai_tables["xai_recon_check"].to_excel(xw, sheet_name="xai_recon_check", index=False)

        pd.DataFrame([asdict(params)]).to_excel(xw, sheet_name="params", index=False)

    print(f"OK: saved {out_path.resolve()}")


if __name__ == "__main__":
    main()
