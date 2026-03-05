from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from pandas.api.types import DatetimeTZDtype
from tqdm.auto import tqdm

from src.v2.api.services.business_reranking import (  # type: ignore
    Candidate,
    RerankParams,
    greedy_rerank,
    business_score,
)

# =========================
# CONFIG
# =========================

RECS_PATH = Path("recommendations_output.parquet")
OUT_XLSX_BASENAME = "compare_top3"

TOPK = 3
P_H = 0.0

API_MODEL_SCORE_MISSING_SENTINEL = -1.0

DEFAULT_CONTRACT_TYPE = "flat"
DEFAULT_CAP_RATIO = 0.0
DEFAULT_M = 0.0
DEFAULT_V = 0

# --- daily_snapshot.offer_performance_metrics snapshot inputs ---
OFFERS_OPM_DATA_PATH = Path("daily_snapshot_offer_performance_metrics_data_20260120_174227.xlsx")
OFFERS_OPM_META_PATH = Path(
    "daily_snapshot_offer_performance_metrics_meta_20260120_174227.xlsx"
)  # optional (diagnostics only)

OFFERS_OPM_DAY_SHIFT_DAYS = 0  # int
OFFERS_OPM_ASOF_DIRECTION = "backward"  # {'backward','forward','nearest'}

CONTRACT_TYPE_CPL = "cpl"   # settlement_type==0
CONTRACT_TYPE_FLAT = "flat"  # settlement_type==1 (also DEFAULT_CONTRACT_TYPE)

RERANK_PARAMS_DICT = dict(
    gamma=1.2,
    mu=0.5,
    nu=0.8,
    rho=0.3,
    delta=1.0,
    lambda_=0.5,
)

# Labels for Excel grouping row (your requirement)
LABEL_RECS = "hm-reranking-testgroup-202512.csv"
LABEL_OPM = "daily_snapshot.offer_performance_metrics"
LABEL_FEATURES = "business reranking: zmienne/cechy (q,r,g,m,v,cap_ratio,...)"
LABEL_WEIGHTS = "business reranking: wagi/parametry (gamma,mu,nu,...)"
LABEL_OUTPUTS = "wyniki rerankingu (rank/score)"
LABEL_OTHER = "inne"

# Background colors per group (header row + column names row)
GROUP_BG = {
    LABEL_RECS: "#DCE6F1",     # light blue
    LABEL_OPM: "#E2EFDA",      # light green
    LABEL_FEATURES: "#FFF2CC", # light yellow
    LABEL_WEIGHTS: "#FCE4D6",  # light orange
    LABEL_OUTPUTS: "#E4DFEC",  # light purple/gray
    LABEL_OTHER: "#F2F2F2",    # light gray
}

# =========================

RECS_REQUIRED_COLS = [
    "create_date",
    "uuid",
    "request_id",
    "offer_id",
    "property_id",
    "api_model_score",
    "final_score",
    "gap_score",
    "lead_price_score",
]

OFFERS_OPM_REQUIRED_COLS = [
    "id",
    "day",
    "offer_id",
    "group_id",
    "monthly_budget",
    "estimated_monthly_budget",
    "group_budget",
    "lead_limit",
    "leads",
    "leads_estimation",
    "estimated_perc_realization",
    "lead_price",
    "estimated_lead_price",
    "settlement_type",
]

# =========================
# HELPERS
# =========================

def read_table(path: Path) -> pd.DataFrame:
    """Generic reader for recs input."""
    suffix = path.suffix.lower()
    if suffix in {".parquet", ".pq"}:
        return pd.read_parquet(path)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    if suffix == ".csv":
        return pd.read_csv(path)
    raise ValueError(f"Unsupported input format: {suffix}")


def ensure_required_columns(df: pd.DataFrame, required: list[str], *, df_name: str = "df") -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"{df_name}: missing required columns: {missing}")


def normalize_create_date(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize create_date to tz-naive Europe/Warsaw (for stable day flooring)."""
    out = df.copy()
    dt = pd.to_datetime(out["create_date"], errors="raise")

    if isinstance(dt.dtype, DatetimeTZDtype):
        dt = dt.dt.tz_convert("Europe/Warsaw").dt.tz_localize(None)

    out["create_date"] = dt
    return out


def read_offer_performance_metrics(path: Path) -> pd.DataFrame:
    """Read daily_snapshot.offer_performance_metrics XLSX dump (sheet: 'data')."""
    if not path.exists():
        raise FileNotFoundError(f"offer_performance_metrics data file not found: {path.resolve()}")

    df = pd.read_excel(path, sheet_name="data")
    ensure_required_columns(df, OFFERS_OPM_REQUIRED_COLS, df_name="offer_performance_metrics")

    out = df.copy()
    out["day"] = pd.to_datetime(out["day"], errors="raise").dt.normalize()

    out["offer_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")
    out["group_id"] = pd.to_numeric(out["group_id"], errors="coerce").astype("Int64")
    out["settlement_type"] = pd.to_numeric(out["settlement_type"], errors="coerce").astype("Int64")

    for c in [
        "monthly_budget",
        "estimated_monthly_budget",
        "group_budget",
        "estimated_perc_realization",
        "lead_price",
        "estimated_lead_price",
    ]:
        out[c] = pd.to_numeric(out[c], errors="coerce").astype(float)

    for c in ["lead_limit", "leads", "leads_estimation"]:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    # Uniqueness check for merge_asof by=offer_id on day
    dup = out.duplicated(["offer_id", "day"], keep=False)
    if dup.any():
        sample = out.loc[dup, ["offer_id", "day"]].head(10)
        raise ValueError(
            "offer_performance_metrics: duplicate rows for (offer_id, day). "
            "This breaks merge_asof assumptions. Sample:\n"
            f"{sample}"
        )

    return out


def enrich_with_offer_performance_metrics(
    recs: pd.DataFrame,
    offers_opm: pd.DataFrame,
) -> pd.DataFrame:
    """Enrich recs with daily_snapshot.offer_performance_metrics and build Candidate business features."""
    if OFFERS_OPM_ASOF_DIRECTION not in {"backward", "forward", "nearest"}:
        raise ValueError(
            "OFFERS_OPM_ASOF_DIRECTION must be one of {'backward','forward','nearest'}, "
            f"got: {OFFERS_OPM_ASOF_DIRECTION!r}"
        )

    out = recs.copy()
    out["offer_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")

    out["rec_day"] = out["create_date"].dt.floor("D") + pd.to_timedelta(
        OFFERS_OPM_DAY_SHIFT_DAYS, unit="D"
    )

    out["_orig_row_order"] = out.index.astype(int)

    left = out.sort_values(["rec_day", "offer_id"], kind="mergesort").reset_index(drop=True)
    right = offers_opm.sort_values(["day", "offer_id"], kind="mergesort").reset_index(drop=True)

    merged = pd.merge_asof(
        left,
        right,
        by="offer_id",
        left_on="rec_day",
        right_on="day",
        direction=OFFERS_OPM_ASOF_DIRECTION,
    )

    merged = (
        merged.sort_values("_orig_row_order", kind="mergesort")
        .drop(columns=["_orig_row_order"])
        .reset_index(drop=True)
    )

    # contract_type from settlement_type
    settlement_num = pd.to_numeric(merged.get("settlement_type"), errors="coerce")
    merged["contract_type"] = DEFAULT_CONTRACT_TYPE
    merged.loc[settlement_num == 0, "contract_type"] = CONTRACT_TYPE_CPL
    merged.loc[settlement_num == 1, "contract_type"] = CONTRACT_TYPE_FLAT

    # inv_id: prefer group_id
    merged["group_id"] = pd.to_numeric(merged.get("group_id"), errors="coerce").astype("Int64")
    merged["inv_id"] = merged["group_id"].fillna(merged["offer_id"]).astype("Int64")

    # m: budget per offer
    budget = pd.to_numeric(merged.get("monthly_budget"), errors="coerce")
    est_budget = pd.to_numeric(merged.get("estimated_monthly_budget"), errors="coerce")
    merged["m"] = budget.fillna(est_budget).fillna(DEFAULT_M).astype(float)

    # v: lead_limit
    merged["v"] = (
        pd.to_numeric(merged.get("lead_limit"), errors="coerce")
        .fillna(0)
        .astype(int)
    )

    # cap_ratio
    cap = pd.to_numeric(merged.get("estimated_perc_realization"), errors="coerce")
    leads = pd.to_numeric(merged.get("leads"), errors="coerce")
    lead_limit = pd.to_numeric(merged.get("lead_limit"), errors="coerce")
    cap_fallback = leads / lead_limit
    cap = cap.fillna(cap_fallback)

    merged["cap_ratio"] = (
        cap.fillna(DEFAULT_CAP_RATIO)
        .clip(lower=0.0, upper=1.0)
        .astype(float)
    )

    total_rows = len(merged)
    matched_rows = int(merged["id"].notna().sum()) if "id" in merged.columns else 0
    pct_matched = matched_rows / total_rows if total_rows else 0.0

    print(
        "offer_performance_metrics merge summary:\n"
        f"  rows_total..................: {total_rows}\n"
        f"  rows_matched_snapshot.......: {matched_rows} ({pct_matched:.1%})\n"
        f"  snapshot_day_min/max........: {offers_opm['day'].min()} / {offers_opm['day'].max()}"
    )

    return merged


def _prepare_q(series: pd.Series) -> pd.Series:
    q = pd.to_numeric(series, errors="raise").fillna(0.0)
    if API_MODEL_SCORE_MISSING_SENTINEL is not None:
        q = q.mask(q == API_MODEL_SCORE_MISSING_SENTINEL, 0.0)
    return q.clip(-1.0, 1.0)


def build_features(df: pd.DataFrame) -> pd.DataFrame:
    """Build all model/reranking features."""
    out = df.copy().reset_index(drop=True)
    out["row_id"] = out.index.astype(int)

    for c in ["api_model_score", "final_score", "gap_score", "lead_price_score"]:
        out[c] = pd.to_numeric(out[c], errors="raise")

    out["q_api"] = _prepare_q(out["api_model_score"])
    out["q_final"] = _prepare_q(out["final_score"])

    out["r"] = out["lead_price_score"].fillna(0.0).clip(0.0, 1.0)
    out["g"] = out["gap_score"].fillna(0.0).clip(0.0, 1.0)

    # Keep pre-filled business features if present
    if "m" in out.columns:
        out["m"] = pd.to_numeric(out["m"], errors="coerce").fillna(DEFAULT_M).astype(float)
    else:
        out["m"] = float(DEFAULT_M)

    if "v" in out.columns:
        out["v"] = pd.to_numeric(out["v"], errors="coerce").fillna(DEFAULT_V).astype(int)
    else:
        out["v"] = int(DEFAULT_V)

    if "contract_type" in out.columns:
        out["contract_type"] = out["contract_type"].fillna(DEFAULT_CONTRACT_TYPE).astype(str)
    else:
        out["contract_type"] = DEFAULT_CONTRACT_TYPE

    if "cap_ratio" in out.columns:
        out["cap_ratio"] = (
            pd.to_numeric(out["cap_ratio"], errors="coerce")
            .fillna(DEFAULT_CAP_RATIO)
            .clip(0.0, 1.0)
        )
    else:
        out["cap_ratio"] = float(DEFAULT_CAP_RATIO)

    if "inv_id" in out.columns:
        out["inv_id"] = pd.to_numeric(out["inv_id"], errors="coerce").astype("Int64")
    else:
        out["inv_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")

    return out


# =========================
# TOP-K
# =========================

def baseline_topk_by(df: pd.DataFrame, k: int) -> pd.DataFrame:
    out = df.sort_values(
        ["request_id", "q_api", "property_id"],
        ascending=[True, False, True],
        kind="mergesort",
    ).copy()

    out["baseline_rank"] = out.groupby("request_id").cumcount() + 1
    return out[out["baseline_rank"] <= k].copy()


def business_topk_by(
    df: pd.DataFrame,
    params: RerankParams,
    k: int,
    p_h: float,
    q_col: str,
    rank_col: str,
    score_col: str,
    desc: str,
) -> pd.DataFrame:
    parts = []
    grouped = df.groupby("request_id", sort=False)

    for _, g in tqdm(grouped, total=grouped.ngroups, desc=desc, unit="request"):
        candidates = []
        for row in g.itertuples(index=False):
            candidates.append(
                Candidate(
                    property_id=int(row.property_id),
                    q=float(getattr(row, q_col)),
                    r=float(row.r),
                    g=float(row.g),
                    m=float(row.m),
                    v=int(row.v),
                    contract_type=row.contract_type,
                    cap_ratio=float(row.cap_ratio),
                    inv_id=int(row.inv_id) if pd.notna(row.inv_id) else None,
                    extra={"row_id": int(row.row_id)},
                )
            )

        ranked = greedy_rerank(candidates, params, k=k, p_h=p_h)

        selected = []
        rows = []
        for rank, cand in enumerate(ranked, start=1):
            score = business_score(cand, params, p_h=p_h, selected=selected)
            selected.append(cand)
            rows.append(
                {
                    "row_id": cand.extra["row_id"],
                    rank_col: rank,
                    score_col: score,
                }
            )

        out = g.merge(pd.DataFrame(rows), on="row_id", how="inner")
        parts.append(out.sort_values(rank_col, kind="mergesort"))

    return pd.concat(parts, ignore_index=True)


# =========================
# SUMMARY
# =========================

def summary_wide_pair(
    left: pd.DataFrame,
    right: pd.DataFrame,
    k: int,
    left_rank: str,
    right_rank: str,
    left_prefix: str,
    right_prefix: str,
    left_score_col: str,
    right_score_col: str,
) -> pd.DataFrame:
    meta = (
        left[["request_id", "uuid", "create_date"]]
        .drop_duplicates("request_id")
        .set_index("request_id")
    )

    l_pid = left.pivot(index="request_id", columns=left_rank, values="property_id")
    l_pid.columns = [f"{left_prefix}_pid_{c}" for c in l_pid.columns]

    l_score = left.pivot(index="request_id", columns=left_rank, values=left_score_col)
    l_score.columns = [f"{left_prefix}_score_{c}" for c in l_score.columns]

    r_pid = right.pivot(index="request_id", columns=right_rank, values="property_id")
    r_pid.columns = [f"{right_prefix}_pid_{c}" for c in r_pid.columns]

    r_score = right.pivot(index="request_id", columns=right_rank, values=right_score_col)
    r_score.columns = [f"{right_prefix}_score_{c}" for c in r_score.columns]

    out = meta.join(l_pid).join(l_score).join(r_pid).join(r_score).reset_index()

    for i in range(1, k + 1):
        for col in (
            f"{left_prefix}_pid_{i}",
            f"{left_prefix}_score_{i}",
            f"{right_prefix}_pid_{i}",
            f"{right_prefix}_score_{i}",
        ):
            if col not in out.columns:
                out[col] = pd.NA

    def overlap(row):
        a = {row[f"{left_prefix}_pid_{i}"] for i in range(1, k + 1)}
        b = {row[f"{right_prefix}_pid_{i}"] for i in range(1, k + 1)}
        return len({x for x in a if pd.notna(x)} & {x for x in b if pd.notna(x)})

    out[f"top{k}_overlap_cnt"] = out.apply(overlap, axis=1)
    return out


# =========================
# METRICS
# =========================

def compute_pair_metrics(
    summary_df: pd.DataFrame,
    k: int,
    left_prefix: str,
    right_prefix: str,
) -> dict:
    n = len(summary_df)

    overlap = summary_df[f"top{k}_overlap_cnt"].fillna(0)

    top1_match = (
        summary_df[f"{left_prefix}_pid_1"]
        == summary_df[f"{right_prefix}_pid_1"]
    ).fillna(False)

    full_match = overlap == k

    return {
        "requests": int(n),
        "avg_overlap_ratio": float((overlap / k).mean()) if n else 0.0,
        "pct_full_topk_match": float(full_match.mean()) if n else 0.0,
        "pct_top1_match": float(top1_match.mean()) if n else 0.0,
    }


def build_metrics_sheet(
    k: int,
    sum_base_pg: pd.DataFrame,
    sum_base_kurekj: pd.DataFrame,
    sum_pg_kurekj: pd.DataFrame,
) -> pd.DataFrame:
    rows = []

    rows.append({
        "comparison": "baseline vs propertygroup",
        **compute_pair_metrics(sum_base_pg, k, "base", "pg"),
    })
    rows.append({
        "comparison": "baseline vs kurekj",
        **compute_pair_metrics(sum_base_kurekj, k, "base", "kurekj"),
    })
    rows.append({
        "comparison": "propertygroup vs kurekj",
        **compute_pair_metrics(sum_pg_kurekj, k, "pg", "kurekj"),
    })

    df = pd.DataFrame(rows)

    df["avg_overlap_ratio_pct"] = (df["avg_overlap_ratio"] * 100).round(2)
    df["pct_full_topk_match_pct"] = (df["pct_full_topk_match"] * 100).round(2)
    df["pct_top1_match_pct"] = (df["pct_top1_match"] * 100).round(2)

    ordered = [
        "comparison",
        "requests",
        "avg_overlap_ratio",
        "avg_overlap_ratio_pct",
        "pct_full_topk_match",
        "pct_full_topk_match_pct",
        "pct_top1_match",
        "pct_top1_match_pct",
    ]

    return df[ordered]


# =========================
# EXCEL HELPERS (GROUP HEADER + COLORS)
# =========================

def choose_excel_engine() -> str:
    try:
        __import__("xlsxwriter")
        return "xlsxwriter"
    except Exception:
        return "openpyxl"


def add_param_columns(df: pd.DataFrame, params: RerankParams, p_h: float) -> pd.DataFrame:
    """Adds constant params columns so they can be grouped/colored in the same sheet."""
    out = df.copy()
    for k, v in asdict(params).items():
        out[f"param_{k}"] = v
    out["p_h"] = p_h
    return out


def _compress_segments(cols: List[str], col2group: Dict[str, str]) -> List[Tuple[str, int, int]]:
    """Return contiguous (group_label, start_col_idx, end_col_idx) segments."""
    if not cols:
        return []
    segs: List[Tuple[str, int, int]] = []
    cur = col2group.get(cols[0], LABEL_OTHER)
    start = 0
    for i in range(1, len(cols)):
        g = col2group.get(cols[i], LABEL_OTHER)
        if g != cur:
            segs.append((cur, start, i - 1))
            cur = g
            start = i
    segs.append((cur, start, len(cols) - 1))
    return segs


def _infer_col_groups(
    df_cols: List[str],
    *,
    recs_input_cols: set[str],
    opm_cols: set[str],
    feature_cols: set[str],
    weight_cols: set[str],
    output_cols: set[str],
) -> Dict[str, str]:
    """Map each column -> group label."""
    col2group: Dict[str, str] = {}
    for c in df_cols:
        if c in output_cols:
            col2group[c] = LABEL_OUTPUTS
        elif c in weight_cols:
            col2group[c] = LABEL_WEIGHTS
        elif c in feature_cols:
            col2group[c] = LABEL_FEATURES
        elif c in opm_cols:
            col2group[c] = LABEL_OPM
        elif c in recs_input_cols:
            col2group[c] = LABEL_RECS
        else:
            col2group[c] = LABEL_OTHER
    return col2group


def write_df_with_group_header(
    xw: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    recs_input_cols: set[str],
    opm_cols: set[str],
    feature_cols: set[str],
    weight_cols: set[str],
    output_cols: set[str],
) -> None:
    """
    Writes df to Excel with:
      - extra row 0: group labels (merged across columns)
      - row 1: normal column headers, but colored per group
      - freeze panes below row 1
    Works for xlsxwriter and openpyxl engines.
    """
    # Write dataframe starting at row=1 (so row=0 is free for group labels)
    df.to_excel(xw, sheet_name=sheet_name, index=False, startrow=1)

    cols = list(df.columns)
    col2group = _infer_col_groups(
        cols,
        recs_input_cols=recs_input_cols,
        opm_cols=opm_cols,
        feature_cols=feature_cols,
        weight_cols=weight_cols,
        output_cols=output_cols,
    )
    segs = _compress_segments(cols, col2group)

    book = xw.book
    ws = xw.sheets[sheet_name]

    is_xlsxwriter = hasattr(book, "add_format") and hasattr(ws, "merge_range")

    if is_xlsxwriter:
        # --- XLSXWRITER PATH ---
        workbook = book

        def fmt(bg: str, *, bold: bool) -> object:
            return workbook.add_format(
                {
                    "bold": bold,
                    "align": "center",
                    "valign": "vcenter",
                    "bg_color": bg,
                    "border": 1,
                    "text_wrap": True,
                }
            )

        group_fmt = {k: fmt(GROUP_BG.get(k, GROUP_BG[LABEL_OTHER]), bold=True) for k in GROUP_BG}
        head_fmt = {k: fmt(GROUP_BG.get(k, GROUP_BG[LABEL_OTHER]), bold=True) for k in GROUP_BG}

        ws.set_row(0, 22)
        ws.set_row(1, 18)

        for g, c0, c1 in segs:
            f = group_fmt.get(g, group_fmt[LABEL_OTHER])
            if c0 == c1:
                ws.write(0, c0, g, f)
            else:
                ws.merge_range(0, c0, 0, c1, g, f)

        for j, c in enumerate(cols):
            g = col2group.get(c, LABEL_OTHER)
            f = head_fmt.get(g, head_fmt[LABEL_OTHER])
            ws.write(1, j, c, f)

        ws.freeze_panes(2, 0)

        if len(cols) > 0:
            ws.autofilter(1, 0, 1, len(cols) - 1)

    else:
        # --- OPENPYXL PATH ---
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        header_row = 2  # because startrow=1 -> pandas header is row 2
        group_row = 1

        ws.row_dimensions[group_row].height = 22
        ws.row_dimensions[header_row].height = 18

        def make_fill(hex_color: str) -> PatternFill:
            c = hex_color.replace("#", "")
            return PatternFill(patternType="solid", fgColor=c)

        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ✅ jawny border (zamiast cell.border = cell.border)
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Group row merges + fill + border
        for g, c0, c1 in segs:
            bg = GROUP_BG.get(g, GROUP_BG[LABEL_OTHER])
            fill = make_fill(bg)

            start_col = c0 + 1
            end_col = c1 + 1

            if start_col != end_col:
                ws.merge_cells(
                    start_row=group_row,
                    start_column=start_col,
                    end_row=group_row,
                    end_column=end_col,
                )

            for col_i in range(start_col, end_col + 1):
                cell = ws.cell(row=group_row, column=col_i)
                cell.fill = fill
                cell.font = bold_font
                cell.alignment = center
                cell.border = border  # ✅ OK

            ws.cell(row=group_row, column=start_col).value = g

        # Header row styles (row=2)
        for j, c in enumerate(cols, start=1):
            g = col2group.get(c, LABEL_OTHER)
            fill = make_fill(GROUP_BG.get(g, GROUP_BG[LABEL_OTHER]))

            cell = ws.cell(row=header_row, column=j)
            cell.value = c
            cell.fill = fill
            cell.font = bold_font
            cell.alignment = center
            cell.border = border  # ✅ OK

            # Optional: set a reasonable width
            col_letter = get_column_letter(j)
            ws.column_dimensions[col_letter].width = max(12, min(45, len(str(c)) + 2))

        # Freeze panes below headers (row 2)
        ws.freeze_panes = ws["A3"]

        # ✅ Autofilter on header row (row 2)
        if len(cols) > 0:
            last_col_letter = get_column_letter(len(cols))
            last_row = df.shape[0] + header_row  # header_row + data_rows
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"


# =========================
# MAIN
# =========================

def main() -> None:
    recs = read_table(RECS_PATH)
    ensure_required_columns(recs, RECS_REQUIRED_COLS, df_name="recs")

    recs = normalize_create_date(recs)

    # Capture input cols (for grouping)
    recs_input_cols_set = set(recs.columns)

    opm_cols_set: set[str] = set()
    if OFFERS_OPM_DATA_PATH.exists():
        offers_opm = read_offer_performance_metrics(OFFERS_OPM_DATA_PATH)
        # columns coming from OPM in merged output (exclude join key offer_id)
        opm_cols_set = set(offers_opm.columns) - {"offer_id"}
        recs = enrich_with_offer_performance_metrics(recs, offers_opm)
    else:
        print(
            "WARN: OFFERS_OPM_DATA_PATH does not exist -> using DEFAULT_* for business features. "
            f"path={OFFERS_OPM_DATA_PATH.resolve()}"
        )

    recs = build_features(recs)

    params = RerankParams(**RERANK_PARAMS_DICT)

    baseline_top = baseline_topk_by(recs, TOPK)

    pg_top = business_topk_by(
        recs,
        params,
        TOPK,
        P_H,
        q_col="q_final",
        rank_col="pg_rank",
        score_col="pg_score",
        desc="Propertygroup reranking (final_score)",
    )

    kurekj_top = business_topk_by(
        recs,
        params,
        TOPK,
        P_H,
        q_col="q_api",
        rank_col="kurekj_rank",
        score_col="kurekj_score",
        desc="Kurekj reranking (api_model_score)",
    )

    # Add constant params columns to topk sheets (so they can be grouped & colored)
    baseline_top = add_param_columns(baseline_top, params, P_H)
    pg_top = add_param_columns(pg_top, params, P_H)
    kurekj_top = add_param_columns(kurekj_top, params, P_H)

    sum_base_pg = summary_wide_pair(
        baseline_top,
        pg_top,
        TOPK,
        "baseline_rank",
        "pg_rank",
        "base",
        "pg",
        "q_api",
        "pg_score",
    )

    sum_base_kurekj = summary_wide_pair(
        baseline_top,
        kurekj_top,
        TOPK,
        "baseline_rank",
        "kurekj_rank",
        "base",
        "kurekj",
        "q_api",
        "kurekj_score",
    )

    sum_pg_kurekj = summary_wide_pair(
        pg_top,
        kurekj_top,
        TOPK,
        "pg_rank",
        "kurekj_rank",
        "pg",
        "kurekj",
        "pg_score",
        "kurekj_score",
    )

    metrics_df = build_metrics_sheet(
        TOPK,
        sum_base_pg,
        sum_base_kurekj,
        sum_pg_kurekj,
    )

    out_dir = Path(__file__).resolve().parent / f"Output_{datetime.now():%Y%m%d_%H%M%S}"
    out_dir.mkdir(parents=True, exist_ok=True)

    out_path = out_dir / f"{OUT_XLSX_BASENAME}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    # Group definitions for Excel (columns sets)
    feature_cols_set = {
        # alignment / merge helper
        "rec_day",
        # created in build_features
        "row_id",
        "q_api",
        "q_final",
        "r",
        "g",
        # business inputs used by rerank
        "m",
        "v",
        "contract_type",
        "cap_ratio",
        "inv_id",
    }
    weight_cols_set = {f"param_{k}" for k in asdict(params).keys()} | {"p_h"}
    output_cols_set = {
        "baseline_rank",
        "pg_rank",
        "pg_score",
        "kurekj_rank",
        "kurekj_score",
    }

    with pd.ExcelWriter(out_path, engine=choose_excel_engine()) as xw:
        # --- TOPK sheets WITH group header row + colors ---
        write_df_with_group_header(
            xw,
            "baseline_topk",
            baseline_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        write_df_with_group_header(
            xw,
            "pg_business_topk",
            pg_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        write_df_with_group_header(
            xw,
            "kurekj_business",
            kurekj_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        # --- The remaining sheets keep the standard single header row (so nic nie psuje analiz/formuł) ---
        sum_base_pg.to_excel(xw, "sum_base_vs_pg", index=False)
        sum_base_kurekj.to_excel(xw, "sum_base_vs_kurekj", index=False)
        sum_pg_kurekj.to_excel(xw, "sum_pg_vs_kurekj", index=False)

        metrics_df.to_excel(xw, "metrics", index=False)
        pd.DataFrame([asdict(params)]).to_excel(xw, "params", index=False)

    print(f"OK: saved {out_path.resolve()}")


if __name__ == "__main__":
    main()
