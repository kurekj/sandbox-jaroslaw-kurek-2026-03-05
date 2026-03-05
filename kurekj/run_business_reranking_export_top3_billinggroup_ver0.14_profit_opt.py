from __future__ import annotations

import os
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from pandas.api.types import DatetimeTZDtype
from tqdm.auto import tqdm

from src.v2.api.services.business_reranking4 import (  # type: ignore
    Candidate,
    RerankParams,
    business_score,
    export_debug_to_excel,
    greedy_rerank,
    rerank_dataframe_debug,
)

# =========================
# CONFIG
# =========================

RECS_PATH = Path("recommendations_output.parquet")
OUT_XLSX_BASENAME = "compare_top3"

TOPK = 3
# P_H = 0.0
P_H = 0.416210088303199

API_MODEL_SCORE_MISSING_SENTINEL = -1.0

DEFAULT_CONTRACT_TYPE = "flat"
DEFAULT_CAP_RATIO = 0.0
DEFAULT_M = 0.0
DEFAULT_V = 0

# -------------------------
# Feature normalization (IMPORTANT)
# -------------------------
# In business_reranking2 the score expects:
#   r,g,cap_ratio in [0,1], m in [-1,1], v in {0,1}.
# This script builds these ranges explicitly (previously m=budget, v=lead_limit).

FAIRNESS_MODE = "budget_norm_log1p"  # {'none','budget_norm_log1p'}
FAIRNESS_BUDGET_BY_CONTRACT = True   # normalize budgets separately for cpl vs flat
FAIRNESS_CLIP = (-1.0, 1.0)

VIP_MODE = "p90_effective_price"     # {'none','binary_from_lead_limit','p90_effective_price'}
VIP_BY_CONTRACT = True              # compute VIP thresholds separately for cpl vs flat

# --- daily_snapshot.offer_performance_metrics snapshot inputs ---
OFFERS_OPM_DATA_PATH = Path("daily_snapshot_offer_performance_metrics_data_20260120_174227.xlsx")
OFFERS_OPM_META_PATH = Path("daily_snapshot_offer_performance_metrics_meta_20260120_174227.xlsx")

OFFERS_OPM_DAY_SHIFT_DAYS = 0  # int
OFFERS_OPM_ASOF_DIRECTION = "backward"  # {'backward','forward','nearest'}

CONTRACT_TYPE_CPL = "cpl"  # settlement_type==0
CONTRACT_TYPE_FLAT = "flat"  # settlement_type==1 (also DEFAULT_CONTRACT_TYPE)

RERANK_PARAMS_DICT = dict(
    gamma=0.19003712932599,
    mu=0.0198973389412471,
    nu=1.99400362001601,
    rho=0.418331944593489,
    delta=0.636301389842897,
    lambda_=1.9534446234665,
)

# --- DEBUG EXPORT (XLSX z business_reranking2) ---
# Uwaga: debug DF może być bardzo szeroki; default to mała próbka requestów.
DEBUG_EXPORT_ENABLED = True
DEBUG_EXPORT_MAX_REQUESTS = 10  # jeśli DEBUG_EXPORT_REQUEST_IDS puste
DEBUG_EXPORT_REQUEST_IDS: List[str] = []  # np. ["daa4d8a3-..."] – wtedy ignoruje MAX_REQUESTS
# ZMIANA: tylko KUREKJ liczymy wg business_score (LaTeX)
DEBUG_EXPORT_ALGOS = ("kurekj",)  # które algorytmy debugować: ("kurekj",) / (puste) / itd.
DEBUG_EXPORT_FILENAME_BASENAME = "debug_breakdown"


# --- EXTRA DEBUG XLSX OUTPUTS (osobne pliki, bez osobnych skryptów) ---
# 1) biznesowy score breakdown TYLKO dla KUREKJ (q_api) – pełny wzór + wkłady (%)
DEBUG_BUSINESS_SCORE_KUREKJ_XLSX_ENABLED = True
DEBUG_BUSINESS_SCORE_KUREKJ_FILENAME_BASENAME = "debug_business_score_kurekj"

# 2) proxy profit/revenue: porównanie PG vs KUREKJ (2 wiersze pod sobą) + dekompozycja różnic
DEBUG_PROFIT_PAIRWISE_XLSX_ENABLED = True
DEBUG_PROFIT_PAIRWISE_FILENAME_BASENAME = "debug_profit_pairwise_pg_vs_kurekj"

# --- Monetization / profit proxy (for choosing pg vs kurekj) ---
POSITION_WEIGHT_MODE = "dcg"  # {'dcg','equal','custom'}
CUSTOM_POSITION_WEIGHTS = {1: 1.0, 2: 0.7, 3: 0.5}  # used only if POSITION_WEIGHT_MODE='custom'

PROFIT_PROB_COL = "q_api"  # {'q_api','q_final'} etc.
BILLABLE_CAP_THRESHOLD = 1.0

LABEL_RECS = "hm-reranking-testgroup-202512.csv"
LABEL_OPM = "daily_snapshot.offer_performance_metrics"
LABEL_FEATURES = "business reranking: zmienne/cechy (q,r,g,m,v,cap_ratio,...)"
LABEL_WEIGHTS = "business reranking: wagi/parametry (gamma,mu,nu,...)"
LABEL_OUTPUTS = "wyniki rerankingu (rank/score)"
LABEL_OTHER = "inne"

GROUP_BG = {
    LABEL_RECS: "#DCE6F1",  # light blue
    LABEL_OPM: "#E2EFDA",  # light green
    LABEL_FEATURES: "#FFF2CC",  # light yellow
    LABEL_WEIGHTS: "#FCE4D6",  # light orange
    LABEL_OUTPUTS: "#E4DFEC",  # light purple/gray
    LABEL_OTHER: "#F2F2F2",  # light gray
}

# =========================

RECS_REQUIRED_COLS = [
    "create_date",
    "uuid",
    "request_id",
    "offer_id",
    "property_id",
    "api_model_score",
    "final_score",
    "gap_score",
    "lead_price_score",
]

OFFERS_OPM_REQUIRED_COLS = [
    "id",
    "day",
    "offer_id",
    "group_id",
    "monthly_budget",
    "estimated_monthly_budget",
    "group_budget",
    "lead_limit",
    "leads",
    "leads_estimation",
    "estimated_perc_realization",
    "lead_price",
    "estimated_lead_price",
    "settlement_type",
]

# =========================
# HELPERS
# =========================


def read_table(path: Path) -> pd.DataFrame:
    """Generic reader for recs input."""
    suffix = path.suffix.lower()
    if suffix in {".parquet", ".pq"}:
        return pd.read_parquet(path)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    if suffix == ".csv":
        return pd.read_csv(path)
    raise ValueError(f"Unsupported input format: {suffix}")


def ensure_required_columns(df: pd.DataFrame, required: list[str], *, df_name: str = "df") -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"{df_name}: missing required columns: {missing}")


def normalize_create_date(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize create_date to tz-naive Europe/Warsaw (for stable day flooring)."""
    out = df.copy()
    dt = pd.to_datetime(out["create_date"], errors="raise")

    if isinstance(dt.dtype, DatetimeTZDtype):
        dt = dt.dt.tz_convert("Europe/Warsaw").dt.tz_localize(None)

    out["create_date"] = dt
    return out


def _col_or_nan(df: pd.DataFrame, col: str) -> pd.Series:
    """Return df[col] if exists else a NaN series (same length)."""
    if col in df.columns:
        return df[col]
    return pd.Series(np.nan, index=df.index)


def robust_minmax_log1p(series: pd.Series, *, by: pd.Series | None = None) -> pd.Series:
    """
    Robust-ish normalization to [0,1] using log1p + min/max within group.

    Why:
      - budgets/prices are heavy-tailed; log1p stabilizes
      - per-contract normalization avoids CPL vs flat scale mismatch

    Returns:
      float series in [0,1] (NaN -> 0).
    """
    x = pd.to_numeric(series, errors="coerce").astype(float)
    x = np.log1p(x.clip(lower=0.0))

    if by is None:
        lo = float(np.nanmin(x.values)) if np.isfinite(x.values).any() else 0.0
        hi = float(np.nanmax(x.values)) if np.isfinite(x.values).any() else 1.0
        denom = hi - lo if hi > lo else 1.0
        out = (x - lo) / denom
        return out.fillna(0.0).clip(0.0, 1.0)

    out = pd.Series(0.0, index=x.index, dtype=float)
    by_key = by.astype(str).fillna("")

    for k, idx in by_key.groupby(by_key).groups.items():
        xs = x.loc[idx]
        finite = np.isfinite(xs.values)
        if not finite.any():
            out.loc[idx] = 0.0
            continue
        lo = float(np.nanmin(xs.values))
        hi = float(np.nanmax(xs.values))
        denom = hi - lo if hi > lo else 1.0
        out.loc[idx] = ((xs - lo) / denom).fillna(0.0).clip(0.0, 1.0)

    return out


def compute_fairness_from_budget(df: pd.DataFrame) -> pd.Series:
    """
    Build m in [-1,1] from budget.

    Default: log1p(budget) -> [0,1] normalized (optionally per contract),
    then center to [-1,1] via 2*x - 1.

    Interpretation:
      - m ~ +1 => high relative budget (within contract segment)
      - m ~ -1 => low relative budget
    """
    if FAIRNESS_MODE == "none":
        return pd.Series(0.0, index=df.index, dtype=float)

    budget = pd.to_numeric(_col_or_nan(df, "m_budget_raw"), errors="coerce").astype(float)

    by = None
    if FAIRNESS_BUDGET_BY_CONTRACT:
        by = _col_or_nan(df, "contract_type").astype(str).str.lower()

    norm01 = robust_minmax_log1p(budget, by=by)
    m = 2.0 * norm01 - 1.0

    lo, hi = FAIRNESS_CLIP
    return m.clip(lo, hi).fillna(0.0).astype(float)


def compute_effective_lead_price_for_features(df: pd.DataFrame) -> pd.Series:
    """
    The same 'unified price per lead' idea as profit analysis,
    but used here to build VIP flag.
    """
    lead_price = pd.to_numeric(_col_or_nan(df, "lead_price"), errors="coerce")
    est_lead_price = pd.to_numeric(_col_or_nan(df, "estimated_lead_price"), errors="coerce")

    settlement = pd.to_numeric(_col_or_nan(df, "settlement_type"), errors="coerce")
    contract = _col_or_nan(df, "contract_type").astype(str).str.lower()

    if settlement.notna().any():
        is_cpl = settlement.eq(0)
    else:
        is_cpl = contract.eq(CONTRACT_TYPE_CPL)

    eff = lead_price.where(is_cpl, est_lead_price)

    # Fallbacks (same idea as profit metric)
    lead_limit = pd.to_numeric(_col_or_nan(df, "lead_limit_raw"), errors="coerce").replace({0: np.nan})
    est_budget = pd.to_numeric(_col_or_nan(df, "estimated_monthly_budget"), errors="coerce")
    monthly_budget = pd.to_numeric(_col_or_nan(df, "monthly_budget"), errors="coerce")
    leads_est = pd.to_numeric(_col_or_nan(df, "leads_estimation"), errors="coerce").replace({0: np.nan})

    budget_any = est_budget.fillna(monthly_budget)
    fallback1 = budget_any / lead_limit
    fallback2 = budget_any / leads_est

    eff = eff.fillna(fallback1).fillna(fallback2)

    return pd.to_numeric(eff, errors="coerce").astype(float)


def compute_vip_flag(df: pd.DataFrame) -> pd.Series:
    """
    Build v in {0,1}.

    Default: v=1 if effective_lead_price is in top 10% (p90) within (optional) contract segment.
    Alternative: v=1 if lead_limit_raw > 0 (legacy behavior).
    """
    if VIP_MODE == "none":
        return pd.Series(0, index=df.index, dtype=int)

    if VIP_MODE == "binary_from_lead_limit":
        lead_limit = pd.to_numeric(_col_or_nan(df, "lead_limit_raw"), errors="coerce").fillna(0.0)
        return (lead_limit > 0).astype(int)

    # VIP_MODE == "p90_effective_price"
    price = pd.to_numeric(_col_or_nan(df, "effective_lead_price"), errors="coerce")

    by = None
    if VIP_BY_CONTRACT:
        by = _col_or_nan(df, "contract_type").astype(str).str.lower()

    if by is None:
        thr = float(price.quantile(0.9)) if price.notna().any() else np.inf
        return (price >= thr).fillna(False).astype(int)

    out = pd.Series(0, index=df.index, dtype=int)
    by_key = by.astype(str).fillna("")

    for k, idx in by_key.groupby(by_key).groups.items():
        p = price.loc[idx]
        thr = float(p.quantile(0.9)) if p.notna().any() else np.inf
        out.loc[idx] = (p >= thr).fillna(False).astype(int)

    return out


def read_offer_performance_metrics(path: Path) -> pd.DataFrame:
    """Read daily_snapshot.offer_performance_metrics XLSX dump (sheet: 'data')."""
    if not path.exists():
        raise FileNotFoundError(f"offer_performance_metrics data file not found: {path.resolve()}")

    df = pd.read_excel(path, sheet_name="data")
    ensure_required_columns(df, OFFERS_OPM_REQUIRED_COLS, df_name="offer_performance_metrics")

    out = df.copy()
    out["day"] = pd.to_datetime(out["day"], errors="raise").dt.normalize()

    out["offer_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")
    out["group_id"] = pd.to_numeric(out["group_id"], errors="coerce").astype("Int64")
    out["settlement_type"] = pd.to_numeric(out["settlement_type"], errors="coerce").astype("Int64")

    for c in [
        "monthly_budget",
        "estimated_monthly_budget",
        "group_budget",
        "estimated_perc_realization",
        "lead_price",
        "estimated_lead_price",
    ]:
        out[c] = pd.to_numeric(out[c], errors="coerce").astype(float)

    for c in ["lead_limit", "leads", "leads_estimation"]:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    # Uniqueness check for merge_asof by=offer_id on day
    dup = out.duplicated(["offer_id", "day"], keep=False)
    if dup.any():
        sample = out.loc[dup, ["offer_id", "day"]].head(10)
        raise ValueError(
            "offer_performance_metrics: duplicate rows for (offer_id, day). "
            "This breaks merge_asof assumptions. Sample:\n"
            f"{sample}"
        )

    return out


def enrich_with_offer_performance_metrics(
    recs: pd.DataFrame,
    offers_opm: pd.DataFrame,
) -> pd.DataFrame:
    """Enrich recs with daily_snapshot.offer_performance_metrics and build Candidate business features."""
    if OFFERS_OPM_ASOF_DIRECTION not in {"backward", "forward", "nearest"}:
        raise ValueError(
            "OFFERS_OPM_ASOF_DIRECTION must be one of {'backward','forward','nearest'}, "
            f"got: {OFFERS_OPM_ASOF_DIRECTION!r}"
        )

    out = recs.copy()
    out["offer_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")

    out["rec_day"] = out["create_date"].dt.floor("D") + pd.to_timedelta(
        OFFERS_OPM_DAY_SHIFT_DAYS, unit="D"
    )

    out["_orig_row_order"] = out.index.astype(int)

    left = out.sort_values(["rec_day", "offer_id"], kind="mergesort").reset_index(drop=True)
    right = offers_opm.sort_values(["day", "offer_id"], kind="mergesort").reset_index(drop=True)

    merged = pd.merge_asof(
        left,
        right,
        by="offer_id",
        left_on="rec_day",
        right_on="day",
        direction=OFFERS_OPM_ASOF_DIRECTION,
    )

    merged = (
        merged.sort_values("_orig_row_order", kind="mergesort")
        .drop(columns=["_orig_row_order"])
        .reset_index(drop=True)
    )

    # contract_type from settlement_type
    settlement_num = pd.to_numeric(merged.get("settlement_type"), errors="coerce")
    merged["contract_type"] = DEFAULT_CONTRACT_TYPE
    merged.loc[settlement_num == 0, "contract_type"] = CONTRACT_TYPE_CPL
    merged.loc[settlement_num == 1, "contract_type"] = CONTRACT_TYPE_FLAT

    # inv_id: prefer group_id
    merged["group_id"] = pd.to_numeric(merged.get("group_id"), errors="coerce").astype("Int64")
    merged["inv_id"] = merged["group_id"].fillna(merged["offer_id"]).astype("Int64")

    # m_budget_raw: budget per offer (raw, not normalized)
    budget = pd.to_numeric(merged.get("monthly_budget"), errors="coerce")
    est_budget = pd.to_numeric(merged.get("estimated_monthly_budget"), errors="coerce")
    merged["m"] = budget.fillna(est_budget).fillna(DEFAULT_M).astype(float)

    # v_raw: lead_limit (raw, not VIP flag)
    merged["v"] = (
        pd.to_numeric(merged.get("lead_limit"), errors="coerce")
        .fillna(0)
        .astype(int)
    )

    # cap_ratio (keep >1 to preserve "over cap" info; scoring saturates anyway)
    cap = pd.to_numeric(merged.get("estimated_perc_realization"), errors="coerce")
    leads = pd.to_numeric(merged.get("leads"), errors="coerce")
    lead_limit = pd.to_numeric(merged.get("lead_limit"), errors="coerce").replace({0: np.nan})
    cap_fallback = leads / lead_limit
    cap = cap.fillna(cap_fallback)

    merged["cap_ratio"] = (
        cap.fillna(DEFAULT_CAP_RATIO)
        .clip(lower=0.0)
        .astype(float)
    )

    total_rows = len(merged)
    matched_rows = int(merged["id"].notna().sum()) if "id" in merged.columns else 0
    pct_matched = matched_rows / total_rows if total_rows else 0.0

    print(
        "offer_performance_metrics merge summary:\n"
        f"  rows_total..................: {total_rows}\n"
        f"  rows_matched_snapshot.......: {matched_rows} ({pct_matched:.1%})\n"
        f"  snapshot_day_min/max........: {offers_opm['day'].min()} / {offers_opm['day'].max()}"
    )

    return merged


def _prepare_q(series: pd.Series) -> pd.Series:
    q = pd.to_numeric(series, errors="raise").fillna(0.0)
    if API_MODEL_SCORE_MISSING_SENTINEL is not None:
        q = q.mask(q == API_MODEL_SCORE_MISSING_SENTINEL, 0.0)
    return q.clip(-1.0, 1.0)


def build_features(df: pd.DataFrame) -> pd.DataFrame:
    """Build all model/reranking features (with correct ranges for business_score)."""
    out = df.copy().reset_index(drop=True)
    out["row_id"] = out.index.astype(int)

    for c in ["api_model_score", "final_score", "gap_score", "lead_price_score"]:
        out[c] = pd.to_numeric(out[c], errors="raise")

    out["q_api"] = _prepare_q(out["api_model_score"])
    out["q_final"] = _prepare_q(out["final_score"])

    # r,g are already 0..1 in logs (scores), but clip to be safe
    out["r"] = out["lead_price_score"].fillna(0.0).clip(0.0, 1.0)
    out["g"] = out["gap_score"].fillna(0.0).clip(0.0, 1.0)

    # contract_type
    if "contract_type" in out.columns:
        out["contract_type"] = out["contract_type"].fillna(DEFAULT_CONTRACT_TYPE).astype(str)
    else:
        out["contract_type"] = DEFAULT_CONTRACT_TYPE

    # cap_ratio (keep >1 if present; only clip lower bound)
    if "cap_ratio" in out.columns:
        out["cap_ratio"] = pd.to_numeric(out["cap_ratio"], errors="coerce").fillna(DEFAULT_CAP_RATIO).clip(lower=0.0)
    else:
        out["cap_ratio"] = float(DEFAULT_CAP_RATIO)

    # inv_id (billing group / investment context for diversity)
    if "inv_id" in out.columns:
        out["inv_id"] = pd.to_numeric(out["inv_id"], errors="coerce").astype("Int64")
    else:
        out["inv_id"] = pd.to_numeric(out["offer_id"], errors="raise").astype("Int64")

    # Preserve raw business inputs if present (from OPM merge)
    if "m" in out.columns:
        out["m_budget_raw"] = pd.to_numeric(out["m"], errors="coerce").fillna(0.0).astype(float)
    else:
        out["m_budget_raw"] = 0.0

    if "v" in out.columns:
        out["lead_limit_raw"] = pd.to_numeric(out["v"], errors="coerce").fillna(0).astype(int)
    else:
        out["lead_limit_raw"] = 0

    # For VIP computation we want effective lead price; store it for debug/profit alignment
    out["effective_lead_price"] = compute_effective_lead_price_for_features(out)

    # m in [-1,1]
    out["m"] = compute_fairness_from_budget(out)

    # v in {0,1}
    out["v"] = compute_vip_flag(out)

    return out


# =========================
# TOP-K
# =========================


def baseline_topk_by(df: pd.DataFrame, k: int) -> pd.DataFrame:
    """
    BASELINE = PROVIDED:
      - api_model_score to dostarczony finalny biznesowy score/ranking
      - NIE wolno liczyć business_score() dla baseline
    """
    out = df.sort_values(
        ["request_id", "api_model_score", "property_id"],
        ascending=[True, False, True],
        kind="mergesort",
    ).copy()

    out["baseline_rank"] = out.groupby("request_id").cumcount() + 1
    return out[out["baseline_rank"] <= k].copy()


def pg_topk_by_final_score(df: pd.DataFrame, k: int) -> pd.DataFrame:
    """
    PG (propertygroup) = PROVIDED:
      - final_score to dostarczony finalny biznesowy score/ranking
      - NIE wolno liczyć business_score() dla PG
    """
    out = df.sort_values(
        ["request_id", "final_score", "property_id"],
        ascending=[True, False, True],
        kind="mergesort",
    ).copy()

    out["pg_rank"] = out.groupby("request_id").cumcount() + 1

    # Ujednolicamy nazwę używaną downstream w arkuszach "sum_*"
    # IMPORTANT: pg_score == dostarczony final_score (bez przeliczeń wzorem)
    out["pg_score"] = pd.to_numeric(out["final_score"], errors="coerce")

    return out[out["pg_rank"] <= k].copy()


def business_topk_by(
    df: pd.DataFrame,
    params: RerankParams,
    k: int,
    p_h: float,
    q_col: str,
    rank_col: str,
    score_col: str,
    desc: str,
) -> pd.DataFrame:
    parts = []
    grouped = df.groupby("request_id", sort=False)

    for _, g in tqdm(grouped, total=grouped.ngroups, desc=desc, unit="request"):
        candidates = []
        for row in g.itertuples(index=False):
            candidates.append(
                Candidate(
                    property_id=int(row.property_id),
                    q=float(getattr(row, q_col)),
                    r=float(row.r),
                    g=float(row.g),
                    m=float(row.m),
                    v=int(row.v),
                    contract_type=row.contract_type,
                    cap_ratio=float(row.cap_ratio),
                    inv_id=int(row.inv_id) if pd.notna(row.inv_id) else None,
                    extra={"row_id": int(row.row_id)},
                )
            )

        ranked = greedy_rerank(candidates, params, k=k, p_h=p_h)

        selected = []
        rows = []
        for rank, cand in enumerate(ranked, start=1):
            score = business_score(cand, params, p_h=p_h, selected=selected)
            selected.append(cand)
            rows.append(
                {
                    "row_id": cand.extra["row_id"],
                    rank_col: rank,
                    score_col: score,
                }
            )

        out = g.merge(pd.DataFrame(rows), on="row_id", how="inner")
        parts.append(out.sort_values(rank_col, kind="mergesort"))

    return pd.concat(parts, ignore_index=True)


# =========================
# SUMMARY
# =========================


def summary_wide_pair(
    left: pd.DataFrame,
    right: pd.DataFrame,
    k: int,
    left_rank: str,
    right_rank: str,
    left_prefix: str,
    right_prefix: str,
    left_score_col: str,
    right_score_col: str,
) -> pd.DataFrame:
    meta = (
        left[["request_id", "uuid", "create_date"]]
        .drop_duplicates("request_id")
        .set_index("request_id")
    )

    l_pid = left.pivot(index="request_id", columns=left_rank, values="property_id")
    l_pid.columns = [f"{left_prefix}_pid_{c}" for c in l_pid.columns]

    l_score = left.pivot(index="request_id", columns=left_rank, values=left_score_col)
    l_score.columns = [f"{left_prefix}_score_{c}" for c in l_score.columns]

    r_pid = right.pivot(index="request_id", columns=right_rank, values="property_id")
    r_pid.columns = [f"{right_prefix}_pid_{c}" for c in r_pid.columns]

    r_score = right.pivot(index="request_id", columns=right_rank, values=right_score_col)
    r_score.columns = [f"{right_prefix}_score_{c}" for c in r_score.columns]

    out = meta.join(l_pid).join(l_score).join(r_pid).join(r_score).reset_index()

    for i in range(1, k + 1):
        for col in (
            f"{left_prefix}_pid_{i}",
            f"{left_prefix}_score_{i}",
            f"{right_prefix}_pid_{i}",
            f"{right_prefix}_score_{i}",
        ):
            if col not in out.columns:
                out[col] = pd.NA

    def overlap(row):
        a = {row[f"{left_prefix}_pid_{i}"] for i in range(1, k + 1)}
        b = {row[f"{right_prefix}_pid_{i}"] for i in range(1, k + 1)}
        return len({x for x in a if pd.notna(x)} & {x for x in b if pd.notna(x)})

    out[f"top{k}_overlap_cnt"] = out.apply(overlap, axis=1)
    return out


# =========================
# METRICS
# =========================


def compute_pair_metrics(
    summary_df: pd.DataFrame,
    k: int,
    left_prefix: str,
    right_prefix: str,
) -> dict:
    n = len(summary_df)

    overlap = summary_df[f"top{k}_overlap_cnt"].fillna(0)

    top1_match = (
        summary_df[f"{left_prefix}_pid_1"]
        == summary_df[f"{right_prefix}_pid_1"]
    ).fillna(False)

    full_match = overlap == k

    return {
        "requests": int(n),
        "avg_overlap_ratio": float((overlap / k).mean()) if n else 0.0,
        "pct_full_topk_match": float(full_match.mean()) if n else 0.0,
        "pct_top1_match": float(top1_match.mean()) if n else 0.0,
    }


def build_metrics_sheet(
    k: int,
    sum_base_pg: pd.DataFrame,
    sum_base_kurekj: pd.DataFrame,
    sum_pg_kurekj: pd.DataFrame,
) -> pd.DataFrame:
    rows = []

    rows.append({
        "comparison": "baseline vs propertygroup",
        **compute_pair_metrics(sum_base_pg, k, "base", "pg"),
    })
    rows.append({
        "comparison": "baseline vs kurekj",
        **compute_pair_metrics(sum_base_kurekj, k, "base", "kurekj"),
    })
    rows.append({
        "comparison": "propertygroup vs kurekj",
        **compute_pair_metrics(sum_pg_kurekj, k, "pg", "kurekj"),
    })

    df = pd.DataFrame(rows)

    df["avg_overlap_ratio_pct"] = (df["avg_overlap_ratio"] * 100).round(2)
    df["pct_full_topk_match_pct"] = (df["pct_full_topk_match"] * 100).round(2)
    df["pct_top1_match_pct"] = (df["pct_top1_match"] * 100).round(2)

    ordered = [
        "comparison",
        "requests",
        "avg_overlap_ratio",
        "avg_overlap_ratio_pct",
        "pct_full_topk_match",
        "pct_full_topk_match_pct",
        "pct_top1_match",
        "pct_top1_match_pct",
    ]

    return df[ordered]


# =========================
# PROFIT / EARNINGS PROXY (PG vs KUREKJ)
# =========================


def _series_or_nan(df: pd.DataFrame, col: str) -> pd.Series:
    """Return df[col] if exists else a NaN series (same length)."""
    if col in df.columns:
        return df[col]
    return pd.Series(np.nan, index=df.index)


def get_position_weights(k: int) -> Dict[int, float]:
    """Position discount weights for TOP-K aggregation."""
    if POSITION_WEIGHT_MODE == "equal":
        return {i: 1.0 for i in range(1, k + 1)}

    if POSITION_WEIGHT_MODE == "custom":
        return {i: float(CUSTOM_POSITION_WEIGHTS.get(i, 0.0)) for i in range(1, k + 1)}

    # default: DCG-like discount
    return {i: float(1.0 / np.log2(i + 1)) for i in range(1, k + 1)}


def compute_effective_lead_price(df: pd.DataFrame) -> pd.Series:
    """
    One unified 'price per lead' to compare earnings across settlement models:
      - CPL: lead_price
      - Subscription/flat: estimated_lead_price (fallbacks if missing)

    Returns float series (may contain NaN if nothing can be computed).
    """
    lead_price = pd.to_numeric(_series_or_nan(df, "lead_price"), errors="coerce")
    est_lead_price = pd.to_numeric(_series_or_nan(df, "estimated_lead_price"), errors="coerce")

    settlement = pd.to_numeric(_series_or_nan(df, "settlement_type"), errors="coerce")
    contract = _series_or_nan(df, "contract_type").astype(str).str.lower()

    # Prefer settlement_type if it has ANY non-null values, otherwise fall back to contract_type
    if settlement.notna().any():
        is_cpl = settlement.eq(0)
    else:
        is_cpl = contract.eq(CONTRACT_TYPE_CPL)

    eff = lead_price.where(is_cpl, est_lead_price)

    # Fallbacks for subscription rows where estimated_lead_price is missing (few % of cases)
    lead_limit = pd.to_numeric(_series_or_nan(df, "lead_limit"), errors="coerce").replace({0: np.nan})
    est_budget = pd.to_numeric(_series_or_nan(df, "estimated_monthly_budget"), errors="coerce")
    monthly_budget = pd.to_numeric(_series_or_nan(df, "monthly_budget"), errors="coerce")
    leads_est = pd.to_numeric(_series_or_nan(df, "leads_estimation"), errors="coerce").replace({0: np.nan})

    budget_any = est_budget.fillna(monthly_budget)

    fallback1 = budget_any / lead_limit
    fallback2 = budget_any / leads_est

    eff = eff.fillna(fallback1).fillna(fallback2)

    return eff.astype(float)


def prepare_profit_frame(
    topk_df: pd.DataFrame,
    *,
    algo: str,
    rank_col: str,
    k: int,
    prob_col: str = PROFIT_PROB_COL,
) -> pd.DataFrame:
    """
    Add monetization columns to a TOP-K dataframe for aggregated 'earnings' comparison.

    Expected revenue proxy per row:
      effective_lead_price * prob(q) * position_weight * billable_flag
    """
    out = topk_df.copy()

    pos_w = get_position_weights(k)
    out["algo"] = algo
    out["rank"] = pd.to_numeric(out[rank_col], errors="coerce").astype("Int64")

    # unify price per lead
    out["effective_lead_price"] = compute_effective_lead_price(out)

    # position weight
    out["position_weight"] = (
        out["rank"].astype(float).map(pos_w).fillna(0.0).astype(float)
    )

    # conversion proxy
    q = pd.to_numeric(_series_or_nan(out, prob_col), errors="coerce").fillna(0.0)
    out["q_prob"] = q.clip(0.0, 1.0)

    # billable flag based on cap_ratio (>=1 => no incremental revenue)
    cap_ratio = pd.to_numeric(_series_or_nan(out, "cap_ratio"), errors="coerce").fillna(0.0)
    out["is_billable"] = (cap_ratio < BILLABLE_CAP_THRESHOLD).astype(int)

    # revenue proxies
    price = out["effective_lead_price"].fillna(0.0).clip(lower=0.0)
    out["unit_value"] = price
    out["unit_value_pos"] = price * out["position_weight"] * out["is_billable"]
    out["expected_revenue"] = price * out["q_prob"] * out["position_weight"] * out["is_billable"]

    return out


def profit_summary_row(df: pd.DataFrame) -> Dict[str, object]:
    reqs = int(df["request_id"].nunique()) if "request_id" in df.columns else 0
    rows = int(len(df))

    eff_price = pd.to_numeric(_series_or_nan(df, "effective_lead_price"), errors="coerce")
    missing_price_pct = float(eff_price.isna().mean()) if rows else 0.0

    out = {
        "algo": str(df["algo"].iloc[0]) if rows and "algo" in df.columns else "",
        "requests": reqs,
        "rows_topk": rows,
        "avg_rows_per_request": float(rows / reqs) if reqs else 0.0,
        "sum_unit_value_pos": float(
            pd.to_numeric(_series_or_nan(df, "unit_value_pos"), errors="coerce").fillna(0.0).sum()),
        "sum_expected_revenue": float(
            pd.to_numeric(_series_or_nan(df, "expected_revenue"), errors="coerce").fillna(0.0).sum()),
        "avg_expected_revenue_per_request": float(
            pd.to_numeric(_series_or_nan(df, "expected_revenue"), errors="coerce").fillna(
                0.0).sum() / reqs) if reqs else 0.0,
        "pct_missing_effective_lead_price": round(missing_price_pct * 100, 2),
        "pct_billable_rows": round(
            float(pd.to_numeric(_series_or_nan(df, "is_billable"), errors="coerce").fillna(0).mean()) * 100,
            2) if rows else 0.0,
        "pct_cpl_rows": round(
            float((_series_or_nan(df, "contract_type").astype(str).str.lower() == CONTRACT_TYPE_CPL).mean()) * 100,
            2) if rows else 0.0,
        "pct_flat_rows": round(
            float((_series_or_nan(df, "contract_type").astype(str).str.lower() == CONTRACT_TYPE_FLAT).mean()) * 100,
            2) if rows else 0.0,
    }
    return out


def build_profit_tables(
    *,
    baseline_top: pd.DataFrame,
    pg_top: pd.DataFrame,
    kurekj_top: pd.DataFrame,
    k: int,
) -> Dict[str, pd.DataFrame]:
    """Build all profit comparison tables for the Excel report."""

    # Prepare enriched frames (do NOT mutate original)
    base_p = prepare_profit_frame(baseline_top, algo="baseline", rank_col="baseline_rank", k=k)
    pg_p = prepare_profit_frame(pg_top, algo="pg", rank_col="pg_rank", k=k)
    kurekj_p = prepare_profit_frame(kurekj_top, algo="kurekj", rank_col="kurekj_rank", k=k)

    # --- Overall summary ---
    overall = pd.DataFrame(
        [
            profit_summary_row(base_p),
            profit_summary_row(pg_p),
            profit_summary_row(kurekj_p),
        ]
    )

    # --- Pairwise PG vs KUREKJ (per-request totals) ---
    pg_req = pg_p.groupby("request_id", as_index=True)["expected_revenue"].sum()
    k_req = kurekj_p.groupby("request_id", as_index=True)["expected_revenue"].sum()

    pair = (
        pd.DataFrame({"pg_expected_revenue": pg_req, "kurekj_expected_revenue": k_req})
        .fillna(0.0)
        .reset_index()
    )
    pair["diff_pg_minus_kurekj"] = pair["pg_expected_revenue"] - pair["kurekj_expected_revenue"]

    n = len(pair)
    pct_pg_win = float((pair["diff_pg_minus_kurekj"] > 0).mean()) if n else 0.0
    pct_k_win = float((pair["diff_pg_minus_kurekj"] < 0).mean()) if n else 0.0
    pct_tie = float((pair["diff_pg_minus_kurekj"] == 0).mean()) if n else 0.0

    pairwise = pd.DataFrame(
        [
            {
                "metric": "requests_compared",
                "value": int(n),
            },
            {
                "metric": "pct_requests_pg_higher",
                "value": round(pct_pg_win * 100, 2),
            },
            {
                "metric": "pct_requests_kurekj_higher",
                "value": round(pct_k_win * 100, 2),
            },
            {
                "metric": "pct_requests_equal",
                "value": round(pct_tie * 100, 2),
            },
            {
                "metric": "avg_diff_pg_minus_kurekj_per_request",
                "value": float(pair["diff_pg_minus_kurekj"].mean()) if n else 0.0,
            },
            {
                "metric": "median_diff_pg_minus_kurekj_per_request",
                "value": float(pair["diff_pg_minus_kurekj"].median()) if n else 0.0,
            },
            {
                "metric": "total_diff_pg_minus_kurekj",
                "value": float(pair["diff_pg_minus_kurekj"].sum()) if n else 0.0,
            },
        ]
    )

    # --- Breakdown by rank (position) ---
    by_rank = (
        pd.concat([base_p, pg_p, kurekj_p], ignore_index=True)
        .groupby(["algo", "rank"], as_index=False)
        .agg(
            rows=("request_id", "count"),
            sum_unit_value_pos=("unit_value_pos", "sum"),
            sum_expected_revenue=("expected_revenue", "sum"),
            avg_expected_revenue=("expected_revenue", "mean"),
        )
        .sort_values(["algo", "rank"], kind="mergesort")
    )

    # --- Breakdown by contract_type ---
    by_contract_src = pd.concat([base_p, pg_p, kurekj_p], ignore_index=True).copy()
    if "contract_type" not in by_contract_src.columns:
        by_contract_src["contract_type"] = ""

    by_contract_src["contract_type"] = by_contract_src["contract_type"].astype(str)

    by_contract = (
        by_contract_src
        .groupby(["algo", "contract_type"], as_index=False)
        .agg(
            rows=("request_id", "count"),
            sum_unit_value_pos=("unit_value_pos", "sum"),
            sum_expected_revenue=("expected_revenue", "sum"),
            avg_expected_revenue=("expected_revenue", "mean"),
        )
        .sort_values(["algo", "contract_type"], kind="mergesort")
    )

    # --- Top billing groups (inv_id) by expected revenue diff (PG - KUREKJ) ---
    pg_inv = pg_p.groupby("inv_id", as_index=True)["expected_revenue"].sum()
    k_inv = kurekj_p.groupby("inv_id", as_index=True)["expected_revenue"].sum()

    inv = (
        pd.DataFrame({"pg_expected_revenue": pg_inv, "kurekj_expected_revenue": k_inv})
        .fillna(0.0)
    )
    inv["diff_pg_minus_kurekj"] = inv["pg_expected_revenue"] - inv["kurekj_expected_revenue"]
    inv["abs_diff"] = inv["diff_pg_minus_kurekj"].abs()

    inv_top = (
        inv.sort_values("abs_diff", ascending=False, kind="mergesort")
        .head(50)
        .reset_index()
        .rename(columns={"inv_id": "billinggroup_or_offer_id"})
        .drop(columns=["abs_diff"])
    )

    return {
        "overall": overall,
        "pairwise": pairwise,
        "by_rank": by_rank,
        "by_contract": by_contract,
        "by_billinggroup_topdiff": inv_top,
    }


def write_df_section(
    xw: pd.ExcelWriter,
    sheet_name: str,
    title: str,
    df: pd.DataFrame,
    *,
    startrow: int,
) -> int:
    """Write a titled dataframe section into a single sheet; returns next startrow."""
    # data starts one row below title
    df.to_excel(xw, sheet_name=sheet_name, index=False, startrow=startrow + 1)

    book = xw.book
    ws = xw.sheets[sheet_name]

    is_xlsxwriter = hasattr(book, "add_format") and hasattr(ws, "write")
    if is_xlsxwriter:
        title_fmt = book.add_format({"bold": True, "font_size": 12})
        ws.write(startrow, 0, title, title_fmt)
    else:
        from openpyxl.styles import Font
        cell = ws.cell(row=startrow + 1, column=1)  # openpyxl is 1-indexed
        cell.value = title
        cell.font = Font(bold=True, size=12)

    # leave one blank row between sections
    return startrow + len(df) + 3


def write_profit_sheet(
    xw: pd.ExcelWriter,
    *,
    sheet_name: str,
    tables: Dict[str, pd.DataFrame],
) -> None:
    """Create one sheet with multiple profit-comparison sections."""
    start = 0
    start = write_df_section(
        xw,
        sheet_name,
        "Overall (expected revenue proxy, using effective_lead_price + q_prob + position_weight)",
        tables["overall"],
        startrow=start,
    )
    start = write_df_section(
        xw,
        sheet_name,
        "PG vs KUREKJ (per-request win rate on expected revenue)",
        tables["pairwise"],
        startrow=start,
    )
    start = write_df_section(
        xw,
        sheet_name,
        "Breakdown by rank (position)",
        tables["by_rank"],
        startrow=start,
    )
    start = write_df_section(
        xw,
        sheet_name,
        "Breakdown by contract_type (cpl vs flat)",
        tables["by_contract"],
        startrow=start,
    )
    _ = write_df_section(
        xw,
        sheet_name,
        "Top billing groups (inv_id) by |diff| in expected revenue (PG - KUREKJ)",
        tables["by_billinggroup_topdiff"],
        startrow=start,
    )


# =========================
# PROFIT DEBUG (PG vs KUREKJ) – pairwise XLSX
# =========================


def _pick_cols(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    """Return df[cols_that_exist] copy (stable order)."""
    keep = [c for c in cols if c in df.columns]
    return df[keep].copy()


def compute_profit_pairwise_diff_decomp(
    pg_p: pd.DataFrame,
    kurekj_p: pd.DataFrame,
    *,
    eps: float = 1e-12,
) -> pd.DataFrame:
    """
    Dekompozycja różnicy PG vs KUREKJ dla proxy-profit na poziomie (request_id, rank).

    Model per-row:
      expected_revenue = effective_lead_price * q_prob * position_weight * is_billable

    Ponieważ to iloczyn, porównanie PG vs KUREKJ robimy log-ilorazami:
      log(ER_pg / ER_k) = log(P_pg/P_k) + log(Q_pg/Q_k) + log(W_pg/W_k) + log(B_pg/B_k)

    Następnie liczymy udział procentowy każdej składowej jako:
      share_x = |log_ratio_x| / sum(|log_ratio_*|) * 100

    To daje intuicyjny "kto wyjaśnia zmianę" między algorytmami.
    """
    cols = [
        "request_id",
        "rank",
        "property_id",
        "offer_id",
        "inv_id",
        "contract_type",
        "cap_ratio",
        "effective_lead_price",
        "q_prob",
        "position_weight",
        "is_billable",
        "expected_revenue",
    ]

    pg = _pick_cols(pg_p, cols).copy()
    k = _pick_cols(kurekj_p, cols).copy()

    # suffixes
    pg = pg.rename(columns={c: f"{c}_pg" for c in pg.columns if c not in {"request_id", "rank"}})
    k = k.rename(columns={c: f"{c}_kurekj" for c in k.columns if c not in {"request_id", "rank"}})

    pair = pg.merge(k, on=["request_id", "rank"], how="outer")

    # fill numeric NaNs with 0 for robust math
    num_cols = [c for c in pair.columns if c.endswith(("_pg", "_kurekj"))]
    for c in num_cols:
        if c in pair.columns:
            pair[c] = pd.to_numeric(pair[c], errors="coerce")

    for c in [
        "effective_lead_price_pg",
        "q_prob_pg",
        "position_weight_pg",
        "is_billable_pg",
        "expected_revenue_pg",
        "cap_ratio_pg",
        "effective_lead_price_kurekj",
        "q_prob_kurekj",
        "position_weight_kurekj",
        "is_billable_kurekj",
        "expected_revenue_kurekj",
        "cap_ratio_kurekj",
    ]:
        if c in pair.columns:
            pair[c] = pair[c].fillna(0.0)

    # log-ratios per factor (eps avoids log(0))
    def log_ratio(a: pd.Series, b: pd.Series) -> pd.Series:
        a = pd.to_numeric(a, errors="coerce").fillna(0.0).astype(float)
        b = pd.to_numeric(b, errors="coerce").fillna(0.0).astype(float)
        return np.log((a + eps) / (b + eps))

    pair["log_ratio_price"] = log_ratio(pair["effective_lead_price_pg"], pair["effective_lead_price_kurekj"])
    pair["log_ratio_qprob"] = log_ratio(pair["q_prob_pg"], pair["q_prob_kurekj"])
    pair["log_ratio_posw"] = log_ratio(pair["position_weight_pg"], pair["position_weight_kurekj"])
    pair["log_ratio_billable"] = log_ratio(pair["is_billable_pg"], pair["is_billable_kurekj"])

    pair["log_ratio_total_by_factors"] = (
        pair["log_ratio_price"]
        + pair["log_ratio_qprob"]
        + pair["log_ratio_posw"]
        + pair["log_ratio_billable"]
    )
    pair["ratio_total_by_factors"] = np.exp(pair["log_ratio_total_by_factors"])

    # Also compute ER log-ratio directly (can differ slightly due to eps / missing)
    pair["er_ratio_pg_over_kurekj"] = (pair["expected_revenue_pg"] + eps) / (pair["expected_revenue_kurekj"] + eps)
    pair["log_ratio_er_direct"] = np.log(pair["er_ratio_pg_over_kurekj"])

    # additive diff
    pair["diff_er_pg_minus_kurekj"] = pair["expected_revenue_pg"] - pair["expected_revenue_kurekj"]

    # Shares (% of abs log-ratio)
    abs_total = (
        pair["log_ratio_price"].abs()
        + pair["log_ratio_qprob"].abs()
        + pair["log_ratio_posw"].abs()
        + pair["log_ratio_billable"].abs()
    )
    abs_total = abs_total.replace({0.0: np.nan})

    pair["share_price_abs_pct"] = (pair["log_ratio_price"].abs() / abs_total * 100.0).fillna(0.0)
    pair["share_qprob_abs_pct"] = (pair["log_ratio_qprob"].abs() / abs_total * 100.0).fillna(0.0)
    pair["share_posw_abs_pct"] = (pair["log_ratio_posw"].abs() / abs_total * 100.0).fillna(0.0)
    pair["share_billable_abs_pct"] = (pair["log_ratio_billable"].abs() / abs_total * 100.0).fillna(0.0)

    # Optional: signed shares (sign tells direction)
    pair["share_price_signed_abs_pct"] = (pair["log_ratio_price"] / abs_total * 100.0).fillna(0.0)
    pair["share_qprob_signed_abs_pct"] = (pair["log_ratio_qprob"] / abs_total * 100.0).fillna(0.0)
    pair["share_posw_signed_abs_pct"] = (pair["log_ratio_posw"] / abs_total * 100.0).fillna(0.0)
    pair["share_billable_signed_abs_pct"] = (pair["log_ratio_billable"] / abs_total * 100.0).fillna(0.0)

    # Order columns (readable)
    ordered = [
        "request_id",
        "rank",
        "property_id_pg",
        "property_id_kurekj",
        "offer_id_pg",
        "offer_id_kurekj",
        "inv_id_pg",
        "inv_id_kurekj",
        "contract_type_pg",
        "contract_type_kurekj",
        "cap_ratio_pg",
        "cap_ratio_kurekj",
        "effective_lead_price_pg",
        "effective_lead_price_kurekj",
        "q_prob_pg",
        "q_prob_kurekj",
        "position_weight_pg",
        "position_weight_kurekj",
        "is_billable_pg",
        "is_billable_kurekj",
        "expected_revenue_pg",
        "expected_revenue_kurekj",
        "diff_er_pg_minus_kurekj",
        "er_ratio_pg_over_kurekj",
        "log_ratio_er_direct",
        "ratio_total_by_factors",
        "log_ratio_total_by_factors",
        "log_ratio_price",
        "log_ratio_qprob",
        "log_ratio_posw",
        "log_ratio_billable",
        "share_price_abs_pct",
        "share_qprob_abs_pct",
        "share_posw_abs_pct",
        "share_billable_abs_pct",
        "share_price_signed_abs_pct",
        "share_qprob_signed_abs_pct",
        "share_posw_signed_abs_pct",
        "share_billable_signed_abs_pct",
    ]
    for c in ordered:
        if c not in pair.columns:
            pair[c] = pd.NA

    return pair[ordered]


def compute_profit_pairwise_request_totals(
    pg_p: pd.DataFrame,
    kurekj_p: pd.DataFrame,
    *,
    k: int,
) -> pd.DataFrame:
    """
    Per-request totals for PG vs KUREKJ (sum over ranks), plus per-rank ER columns.
    """
    # meta
    meta_cols = [c for c in ["request_id", "uuid", "create_date"] if c in pg_p.columns]
    meta = (
        pg_p[meta_cols]
        .drop_duplicates("request_id")
        .set_index("request_id")
        if meta_cols
        else pd.DataFrame(index=pg_p["request_id"].astype(str).unique())
    )

    pg_tot = pg_p.groupby("request_id", as_index=True)["expected_revenue"].sum().rename("pg_expected_revenue_total")
    k_tot = kurekj_p.groupby("request_id", as_index=True)["expected_revenue"].sum().rename("kurekj_expected_revenue_total")

    out = meta.join(pg_tot).join(k_tot).fillna(0.0)
    out["diff_pg_minus_kurekj_total"] = out["pg_expected_revenue_total"] - out["kurekj_expected_revenue_total"]

    # per-rank columns
    pg_rank = pg_p.pivot_table(index="request_id", columns="rank", values="expected_revenue", aggfunc="sum")
    k_rank = kurekj_p.pivot_table(index="request_id", columns="rank", values="expected_revenue", aggfunc="sum")

    for i in range(1, k + 1):
        out[f"pg_er_rank_{i}"] = pg_rank.get(i, np.nan)
        out[f"kurekj_er_rank_{i}"] = k_rank.get(i, np.nan)
        out[f"diff_er_rank_{i}"] = out[f"pg_er_rank_{i}"].fillna(0.0) - out[f"kurekj_er_rank_{i}"].fillna(0.0)

    out = out.reset_index()

    # stable column order
    ordered = ["request_id"] + [c for c in ["uuid", "create_date"] if c in out.columns] + [
        "pg_expected_revenue_total",
        "kurekj_expected_revenue_total",
        "diff_pg_minus_kurekj_total",
    ]
    for i in range(1, k + 1):
        ordered += [f"pg_er_rank_{i}", f"kurekj_er_rank_{i}", f"diff_er_rank_{i}"]

    for c in ordered:
        if c not in out.columns:
            out[c] = pd.NA

    return out[ordered]


def export_profit_pairwise_debug_xlsx(
    *,
    pg_top: pd.DataFrame,
    kurekj_top: pd.DataFrame,
    request_ids: List[str],
    out_path: Path,
    k: int,
) -> Path:
    """
    Zapisuje osobny XLSX do debugowania profit-proxy:
      - rows_pairwise: 2 wiersze na (request_id, rank): KUREKJ i PG
      - diff_decomp: dekompozycja różnicy (log-ratio) per (request_id, rank)
      - request_totals: sumy per request + rozbicie per rank
    """
    if not request_ids:
        raise ValueError("export_profit_pairwise_debug_xlsx: request_ids is empty")

    rid_set = {str(x) for x in request_ids}

    pg_sub = pg_top[pg_top["request_id"].astype(str).isin(rid_set)].copy()
    k_sub = kurekj_top[kurekj_top["request_id"].astype(str).isin(rid_set)].copy()

    pg_p = prepare_profit_frame(pg_sub, algo="pg", rank_col="pg_rank", k=k)
    k_p = prepare_profit_frame(k_sub, algo="kurekj", rank_col="kurekj_rank", k=k)

    # Tidy rows for easy visual comparison (2 rows under each other)
    keep_cols = [
        "request_id",
        "uuid",
        "create_date",
        "algo",
        "rank",
        "property_id",
        "offer_id",
        "inv_id",
        "contract_type",
        "cap_ratio",
        "is_billable",
        "effective_lead_price",
        "q_prob",
        "position_weight",
        "unit_value",
        "unit_value_pos",
        "expected_revenue",
    ]

    rows = pd.concat(
        [_pick_cols(k_p, keep_cols), _pick_cols(pg_p, keep_cols)],
        ignore_index=True,
    )

    # Sort: request_id, rank, algo (kurekj first, then pg)
    rows["algo"] = rows["algo"].astype(str)
    rows["_algo_order"] = pd.Categorical(rows["algo"], categories=["kurekj", "pg"], ordered=True)
    rows = rows.sort_values(["request_id", "rank", "_algo_order"], kind="mergesort").drop(columns=["_algo_order"])

    diff_decomp = compute_profit_pairwise_diff_decomp(pg_p, k_p)
    req_totals = compute_profit_pairwise_request_totals(pg_p, k_p, k=k)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine=choose_excel_engine()) as xw:
        rows.to_excel(xw, sheet_name="rows_pairwise", index=False)
        diff_decomp.to_excel(xw, sheet_name="diff_decomp", index=False)
        req_totals.to_excel(xw, sheet_name="request_totals", index=False)

        # quick notes
        notes = pd.DataFrame(
            [
                {
                    "what": "expected_revenue proxy per row",
                    "formula": "effective_lead_price * q_prob * position_weight * is_billable",
                },
                {
                    "what": "diff_decomp idea",
                    "formula": "log(ER_pg/ER_k) decomposed into log-ratios of each factor",
                },
            ]
        )
        notes.to_excel(xw, sheet_name="notes", index=False)

    return out_path.resolve()


# =========================
# EXCEL HELPERS (GROUP HEADER + COLORS)
# =========================


def choose_excel_engine() -> str:
    try:
        __import__("xlsxwriter")
        return "xlsxwriter"
    except Exception:
        return "openpyxl"


def add_param_columns(df: pd.DataFrame, params: RerankParams, p_h: float) -> pd.DataFrame:
    """Adds constant params columns so they can be grouped/colored in the same sheet."""
    out = df.copy()
    for k, v in asdict(params).items():
        out[f"param_{k}"] = v
    out["p_h"] = p_h
    return out


def _compress_segments(cols: List[str], col2group: Dict[str, str]) -> List[Tuple[str, int, int]]:
    """Return contiguous (group_label, start_col_idx, end_col_idx) segments."""
    if not cols:
        return []
    segs: List[Tuple[str, int, int]] = []
    cur = col2group.get(cols[0], LABEL_OTHER)
    start = 0
    for i in range(1, len(cols)):
        g = col2group.get(cols[i], LABEL_OTHER)
        if g != cur:
            segs.append((cur, start, i - 1))
            cur = g
            start = i
    segs.append((cur, start, len(cols) - 1))
    return segs


def _infer_col_groups(
    df_cols: List[str],
    *,
    recs_input_cols: set[str],
    opm_cols: set[str],
    feature_cols: set[str],
    weight_cols: set[str],
    output_cols: set[str],
) -> Dict[str, str]:
    """Map each column -> group label."""
    col2group: Dict[str, str] = {}
    for c in df_cols:
        if c in output_cols:
            col2group[c] = LABEL_OUTPUTS
        elif c in weight_cols:
            col2group[c] = LABEL_WEIGHTS
        elif c in feature_cols:
            col2group[c] = LABEL_FEATURES
        elif c in opm_cols:
            col2group[c] = LABEL_OPM
        elif c in recs_input_cols:
            col2group[c] = LABEL_RECS
        else:
            col2group[c] = LABEL_OTHER
    return col2group


def write_df_with_group_header(
    xw: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    recs_input_cols: set[str],
    opm_cols: set[str],
    feature_cols: set[str],
    weight_cols: set[str],
    output_cols: set[str],
) -> None:
    """
    Writes df to Excel with:
      - extra row 0: group labels (merged across columns)
      - row 1: normal column headers, but colored per group
      - freeze panes below row 1
    Works for xlsxwriter and openpyxl engines.
    """
    # Write dataframe starting at row=1 (so row=0 is free for group labels)
    df.to_excel(xw, sheet_name=sheet_name, index=False, startrow=1)

    cols = list(df.columns)
    col2group = _infer_col_groups(
        cols,
        recs_input_cols=recs_input_cols,
        opm_cols=opm_cols,
        feature_cols=feature_cols,
        weight_cols=weight_cols,
        output_cols=output_cols,
    )
    segs = _compress_segments(cols, col2group)

    book = xw.book
    ws = xw.sheets[sheet_name]

    is_xlsxwriter = hasattr(book, "add_format") and hasattr(ws, "merge_range")

    if is_xlsxwriter:
        # --- XLSXWRITER PATH ---
        workbook = book

        def fmt(bg: str, *, bold: bool) -> object:
            return workbook.add_format(
                {
                    "bold": bold,
                    "align": "center",
                    "valign": "vcenter",
                    "bg_color": bg,
                    "border": 1,
                    "text_wrap": True,
                }
            )

        group_fmt = {k: fmt(GROUP_BG.get(k, GROUP_BG[LABEL_OTHER]), bold=True) for k in GROUP_BG}
        head_fmt = {k: fmt(GROUP_BG.get(k, GROUP_BG[LABEL_OTHER]), bold=True) for k in GROUP_BG}

        ws.set_row(0, 22)
        ws.set_row(1, 18)

        for g, c0, c1 in segs:
            f = group_fmt.get(g, group_fmt[LABEL_OTHER])
            if c0 == c1:
                ws.write(0, c0, g, f)
            else:
                ws.merge_range(0, c0, 0, c1, g, f)

        for j, c in enumerate(cols):
            g = col2group.get(c, LABEL_OTHER)
            f = head_fmt.get(g, head_fmt[LABEL_OTHER])
            ws.write(1, j, c, f)

        ws.freeze_panes(2, 0)

        if len(cols) > 0:
            ws.autofilter(1, 0, 1, len(cols) - 1)

    else:
        # --- OPENPYXL PATH ---
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        header_row = 2  # because startrow=1 -> pandas header is row 2
        group_row = 1

        ws.row_dimensions[group_row].height = 22
        ws.row_dimensions[header_row].height = 18

        def make_fill(hex_color: str) -> PatternFill:
            c = hex_color.replace("#", "")
            return PatternFill(patternType="solid", fgColor=c)

        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # jawny border
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Group row merges + fill + border
        for g, c0, c1 in segs:
            bg = GROUP_BG.get(g, GROUP_BG[LABEL_OTHER])
            fill = make_fill(bg)

            start_col = c0 + 1
            end_col = c1 + 1

            if start_col != end_col:
                ws.merge_cells(
                    start_row=group_row,
                    start_column=start_col,
                    end_row=group_row,
                    end_column=end_col,
                )

            for col_i in range(start_col, end_col + 1):
                cell = ws.cell(row=group_row, column=col_i)
                cell.fill = fill
                cell.font = bold_font
                cell.alignment = center
                cell.border = border

            ws.cell(row=group_row, column=start_col).value = g

        # Header row styles (row=2)
        for j, c in enumerate(cols, start=1):
            g = col2group.get(c, LABEL_OTHER)
            fill = make_fill(GROUP_BG.get(g, GROUP_BG[LABEL_OTHER]))

            cell = ws.cell(row=header_row, column=j)
            cell.value = c
            cell.fill = fill
            cell.font = bold_font
            cell.alignment = center
            cell.border = border

            # Optional: set a reasonable width
            col_letter = get_column_letter(j)
            ws.column_dimensions[col_letter].width = max(12, min(45, len(str(c)) + 2))

        # Freeze panes below headers (row 2)
        ws.freeze_panes = ws["A3"]

        # Autofilter on header row (row 2)
        if len(cols) > 0:
            last_col_letter = get_column_letter(len(cols))
            last_row = df.shape[0] + header_row  # header_row + data_rows
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"


# =========================
# MAIN
# =========================


def _select_debug_request_ids(recs: pd.DataFrame) -> List[str]:
    if DEBUG_EXPORT_REQUEST_IDS:
        available = set(recs["request_id"].astype(str).unique().tolist())
        chosen = [rid for rid in DEBUG_EXPORT_REQUEST_IDS if str(rid) in available]
        if not chosen:
            print("WARN: DEBUG_EXPORT_REQUEST_IDS set, but none found in input data -> debug skipped.")
        return chosen

    if DEBUG_EXPORT_MAX_REQUESTS and DEBUG_EXPORT_MAX_REQUESTS > 0:
        # Najnowsze requesty (po create_date) – stabilnie i sensownie do debugowania
        return (
            recs.sort_values("create_date", ascending=False)["request_id"]
            .drop_duplicates()
            .head(DEBUG_EXPORT_MAX_REQUESTS)
            .astype(str)
            .tolist()
        )

    return []


def main() -> None:
    recs = read_table(RECS_PATH)
    ensure_required_columns(recs, RECS_REQUIRED_COLS, df_name="recs")

    recs = normalize_create_date(recs)

    # Capture input cols (for grouping)
    recs_input_cols_set = set(recs.columns)

    opm_cols_set: set[str] = set()
    if OFFERS_OPM_DATA_PATH.exists():
        offers_opm = read_offer_performance_metrics(OFFERS_OPM_DATA_PATH)
        # columns coming from OPM in merged output (exclude join key offer_id)
        opm_cols_set = set(offers_opm.columns) - {"offer_id"}
        recs = enrich_with_offer_performance_metrics(recs, offers_opm)
    else:
        print(
            "WARN: OFFERS_OPM_DATA_PATH does not exist -> using DEFAULT_* for business features. "
            f"path={OFFERS_OPM_DATA_PATH.resolve()}"
        )

    recs = build_features(recs)

    params = RerankParams(**RERANK_PARAMS_DICT)

    # BASELINE: provided api_model_score (no business_score)
    baseline_top = baseline_topk_by(recs, TOPK)

    # PG: provided final_score (no business_score)
    pg_top = pg_topk_by_final_score(recs, TOPK)

    # KUREKJ: ONLY algorithm computed by business_score (LaTeX)
    kurekj_top = business_topk_by(
        recs,
        params,
        TOPK,
        P_H,
        q_col="q_api",
        rank_col="kurekj_rank",
        score_col="kurekj_score",
        desc="Kurekj reranking (business_score over api_model_score)",
    )

    # Add constant params columns to topk sheets (so they can be grouped & colored)
    baseline_top = add_param_columns(baseline_top, params, P_H)
    pg_top = add_param_columns(pg_top, params, P_H)
    kurekj_top = add_param_columns(kurekj_top, params, P_H)

    # ZMIANA: baseline score do porównań = api_model_score (dostarczone)
    sum_base_pg = summary_wide_pair(
        baseline_top,
        pg_top,
        TOPK,
        "baseline_rank",
        "pg_rank",
        "base",
        "pg",
        "api_model_score",
        "pg_score",
    )

    sum_base_kurekj = summary_wide_pair(
        baseline_top,
        kurekj_top,
        TOPK,
        "baseline_rank",
        "kurekj_rank",
        "base",
        "kurekj",
        "api_model_score",
        "kurekj_score",
    )

    sum_pg_kurekj = summary_wide_pair(
        pg_top,
        kurekj_top,
        TOPK,
        "pg_rank",
        "kurekj_rank",
        "pg",
        "kurekj",
        "pg_score",
        "kurekj_score",
    )

    metrics_df = build_metrics_sheet(
        TOPK,
        sum_base_pg,
        sum_base_kurekj,
        sum_pg_kurekj,
    )

    # -------------------------
    # OUTPUT DIR (dynamic)
    # -------------------------
    script_path = Path(__file__).resolve()
    script_filename = script_path.name  # np. run_business_reranking_export_top3_billinggroup_ver0.10_profit_opt.py
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    out_dir = script_path.parent / f"Output_[{script_filename}]_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)

    # podaj Output dir do business_reranking2 (debug)
    os.environ["BUSINESS_RERANKING2_OUTPUT_DIR"] = str(out_dir)
    os.environ["BUSINESS_RERANKING2_TS"] = ts

    out_path = out_dir / f"{OUT_XLSX_BASENAME}_{ts}.xlsx"

    # Group definitions for Excel (columns sets)
    feature_cols_set = {
        "rec_day",
        "row_id",
        "q_api",
        "q_final",
        "r",
        "g",
        "m",               # fairness (normalized)
        "v",               # vip flag (0/1)
        "m_budget_raw",    # raw budget used to derive m
        "lead_limit_raw",  # raw lead_limit (legacy)
        "effective_lead_price",  # for VIP + profit alignment
        "contract_type",
        "cap_ratio",
        "inv_id",
    }
    weight_cols_set = {f"param_{k}" for k in asdict(params).keys()} | {"p_h"}
    output_cols_set = {
        "baseline_rank",
        "pg_rank",
        "pg_score",
        "kurekj_rank",
        "kurekj_score",
    }

    # -------------------------
    # MAIN EXCEL
    # -------------------------
    with pd.ExcelWriter(out_path, engine=choose_excel_engine()) as xw:
        write_df_with_group_header(
            xw,
            "baseline_topk",
            baseline_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        # UWAGA: nazwa sheet pozostaje dla kompatybilności (choć PG nie jest liczone business_score)
        write_df_with_group_header(
            xw,
            "pg_business_topk",
            pg_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        write_df_with_group_header(
            xw,
            "kurekj_business",
            kurekj_top,
            recs_input_cols=recs_input_cols_set,
            opm_cols=opm_cols_set,
            feature_cols=feature_cols_set,
            weight_cols=weight_cols_set,
            output_cols=output_cols_set,
        )

        sum_base_pg.to_excel(xw, sheet_name="sum_base_vs_pg", index=False)
        sum_base_kurekj.to_excel(xw, sheet_name="sum_base_vs_kurekj", index=False)
        sum_pg_kurekj.to_excel(xw, sheet_name="sum_pg_vs_kurekj", index=False)

        metrics_df.to_excel(xw, sheet_name="metrics", index=False)

        profit_tables = build_profit_tables(
            baseline_top=baseline_top,
            pg_top=pg_top,
            kurekj_top=kurekj_top,
            k=TOPK,
        )
        write_profit_sheet(
            xw,
            sheet_name="profit_comparison",
            tables=profit_tables,
        )

        pd.DataFrame([asdict(params)]).to_excel(xw, sheet_name="params", index=False)

    print(f"OK: saved {out_path.resolve()}")

    # -------------------------
    # DEBUG EXCEL (business_reranking2)
    # -------------------------
    if DEBUG_EXPORT_ENABLED:
        debug_request_ids = _select_debug_request_ids(recs)
        if not debug_request_ids:
            print("DEBUG: no request_ids selected -> skipping debug export.")
            return

        topk_parts: List[pd.DataFrame] = []
        debug_parts: List[pd.DataFrame] = []

        for algo in DEBUG_EXPORT_ALGOS:
            # ZMIANA: PG traktujemy jako "provided final_score" -> nie debugujemy business_score dla PG
            if algo == "pg":
                print("INFO: PG is treated as provided final_score -> skipping business_score debug export for PG.")
                continue
            elif algo == "kurekj":
                q_col = "q_api"
            else:
                print(f"WARN: unknown debug algo={algo!r} -> skipping")
                continue

            for rid in debug_request_ids:
                g = recs[recs["request_id"].astype(str) == str(rid)].copy()
                if g.empty:
                    continue

                # rerank_dataframe_debug expects ONE request set
                topk_df, dbg_df = rerank_dataframe_debug(
                    g,
                    params,
                    k=TOPK,
                    p_h=P_H,
                    id_col="property_id",
                    score_col=q_col,
                    r_col="r",
                    g_col="g",
                    m_col="m",
                    v_col="v",
                    contract_col="contract_type",
                    cap_ratio_col="cap_ratio",
                    inv_id_col="inv_id",
                    keep_original_cols=True,
                )

                # add meta
                topk_df["debug_algo"] = algo
                dbg_df["debug_algo"] = algo
                dbg_df["request_id"] = str(rid)

                if "uuid" in g.columns:
                    dbg_df["uuid"] = str(g["uuid"].iloc[0])
                if "create_date" in g.columns:
                    dbg_df["create_date"] = g["create_date"].iloc[0]

                topk_parts.append(topk_df)
                debug_parts.append(dbg_df)

        if topk_parts and debug_parts:
            topk_all = pd.concat(topk_parts, ignore_index=True)
            debug_all = pd.concat(debug_parts, ignore_index=True)

            debug_path = export_debug_to_excel(
                topk_all,
                debug_all,
                path=f"{DEBUG_EXPORT_FILENAME_BASENAME}_{ts}.xlsx",
                topk_sheet="topk_breakdown",
                debug_sheet="debug_greedy_all",
            )
            print(f"OK: saved DEBUG XLSX {debug_path}")
            # --- EXTRA 1: osobny XLSX tylko dla KUREKJ (business score breakdown) ---
            if DEBUG_BUSINESS_SCORE_KUREKJ_XLSX_ENABLED and ("kurekj" in DEBUG_EXPORT_ALGOS):
                topk_k = topk_all[topk_all["debug_algo"] == "kurekj"].copy()
                debug_k = debug_all[debug_all["debug_algo"] == "kurekj"].copy()

                if not topk_k.empty and not debug_k.empty:
                    kurekj_path = export_debug_to_excel(
                        topk_k,
                        debug_k,
                        path=out_dir / f"{DEBUG_BUSINESS_SCORE_KUREKJ_FILENAME_BASENAME}_{ts}.xlsx",
                        topk_sheet="topk_breakdown",
                        debug_sheet="debug_greedy_all",
                    )
                    print(f"OK: saved KUREKJ business-score DEBUG XLSX {kurekj_path}")
                else:
                    print("DEBUG: kurekj-only filter empty -> skipping kurekj business-score xlsx.")

        else:
            print("DEBUG: nothing to export (no topk/debug rows) -> skipping debug xlsx.")

        # --- EXTRA 2: PROFIT proxy pairwise PG vs KUREKJ ---
        if DEBUG_PROFIT_PAIRWISE_XLSX_ENABLED:
            try:
                profit_path = export_profit_pairwise_debug_xlsx(
                    pg_top=pg_top,
                    kurekj_top=kurekj_top,
                    request_ids=debug_request_ids,
                    out_path=out_dir / f"{DEBUG_PROFIT_PAIRWISE_FILENAME_BASENAME}_{ts}.xlsx",
                    k=TOPK,
                )
                print(f"OK: saved PROFIT pairwise DEBUG XLSX {profit_path}")
            except Exception as e:
                print(f"WARN: profit pairwise debug export failed: {e}")


if __name__ == "__main__":
    main()
